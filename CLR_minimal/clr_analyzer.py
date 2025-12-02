
"""
CLR (Canonical LR or LR(1)) Parser Analyzer
Most powerful LR parser - uses lookahead symbols in items
"""

from collections import defaultdict, deque
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class LR1Item:
    """Represents an LR(1) item: [A → α·β, a] where a is lookahead"""
    def __init__(self, lhs, rhs, dot_pos, lookahead):
        self.lhs = lhs
        self.rhs = rhs
        self.dot_pos = dot_pos
        self.lookahead = lookahead  # Single terminal symbol
    
    def __eq__(self, other):
        return (self.lhs == other.lhs and 
                self.rhs == other.rhs and 
                self.dot_pos == other.dot_pos and
                self.lookahead == other.lookahead)
    
    def __hash__(self):
        return hash((self.lhs, tuple(self.rhs), self.dot_pos, self.lookahead))
    
    def __repr__(self):
        rhs_with_dot = self.rhs[:self.dot_pos] + ['·'] + self.rhs[self.dot_pos:]
        return f"[{self.lhs} → {' '.join(rhs_with_dot)}, {self.lookahead}]"
    
    def is_complete(self):
        return self.dot_pos >= len(self.rhs)
    
    def next_symbol(self):
        if self.is_complete():
            return None
        return self.rhs[self.dot_pos]
    
    def advance(self):
        return LR1Item(self.lhs, self.rhs, self.dot_pos + 1, self.lookahead)


class LR1State:
    """Represents a state in LR(1) automaton"""
    def __init__(self, state_id, items):
        self.id = state_id
        self.items = frozenset(items)
    
    def __eq__(self, other):
        return self.items == other.items
    
    def __hash__(self):
        return hash(self.items)
    
    def __repr__(self):
        return f"State {self.id}:\n" + "\n".join(f"  {item}" for item in self.items)


class CLRAnalyzer:
    def __init__(self, grammar_text):
        self.grammar = {}
        self.terminals = set()
        self.non_terminals = set()
        self.start_symbol = None
        self.augmented_start = None
        
        self.states = []
        self.state_map = {}
        self.transitions = {}
        
        self.first = defaultdict(set)
        self.action_table = defaultdict(dict)
        self.goto_table = defaultdict(dict)
        self.conflicts = []
        
        self.parse_grammar(grammar_text)
        self.augment_grammar()
        self.compute_first()
        
    def parse_grammar(self, grammar_text):
        """Parse the grammar from text format"""
        lines = grammar_text.strip().split('\n')
        current_lhs = None
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
                
            if '→' in line:
                parts = line.split('→')
                lhs = parts[0].strip()
                rhs = parts[1].strip()
                current_lhs = lhs
                self.non_terminals.add(lhs)
                
                if self.start_symbol is None:
                    self.start_symbol = lhs
                
                if lhs not in self.grammar:
                    self.grammar[lhs] = []
                self.grammar[lhs].append(self.parse_production(rhs))
                
            elif '|' in line:
                rhs = line.split('|')[1].strip()
                if current_lhs:
                    self.grammar[current_lhs].append(self.parse_production(rhs))
        
        # Extract terminals
        for lhs, productions in self.grammar.items():
            for prod in productions:
                for symbol in prod:
                    if symbol not in self.non_terminals:
                        self.terminals.add(symbol)
        
        self.terminals.add('$')
    
    def parse_production(self, rhs):
        """Parse a production right-hand side into symbols"""
        symbols = []
        i = 0
        while i < len(rhs):
            if rhs[i] == "'":
                j = i + 1
                while j < len(rhs) and rhs[j] != "'":
                    j += 1
                symbols.append(rhs[i:j+1])
                i = j + 1
            elif rhs[i].isspace():
                i += 1
            else:
                j = i
                while j < len(rhs) and not rhs[j].isspace() and rhs[j] != "'":
                    j += 1
                symbol = rhs[i:j]
                if symbol:
                    symbols.append(symbol)
                i = j
        
        return symbols
    
    def augment_grammar(self):
        """Augment grammar with SBar → S"""
        self.augmented_start = self.start_symbol + "Bar"
        self.grammar[self.augmented_start] = [[self.start_symbol]]
        self.non_terminals.add(self.augmented_start)
    
    def compute_first(self):
        """Compute FIRST sets for terminals and non-terminals"""
        print("\n=== COMPUTING FIRST SETS ===\n")
        
        # Initialize FIRST for terminals
        for t in self.terminals:
            self.first[t].add(t)
        
        # Compute FIRST for non-terminals
        changed = True
        while changed:
            changed = False
            for A in self.non_terminals:
                for production in self.grammar.get(A, []):
                    if production:
                        for symbol in production:
                            before = len(self.first[A])
                            self.first[A].update(self.first.get(symbol, set()) - {'ε'})
                            if len(self.first[A]) > before:
                                changed = True
                            # If symbol is terminal, stop
                            if symbol in self.terminals:
                                break
                            # For simplicity, assume non-terminals are not nullable
                            # So we stop after first symbol
                            break
        
        for A in sorted(self.non_terminals):
            print(f"FIRST({A}) = {{{', '.join(sorted(self.first[A]))}}}")
    
    def closure(self, items):
        """Compute closure of a set of LR(1) items"""
        closure_set = set(items)
        changed = True
        
        while changed:
            changed = False
            new_items = set()
            
            for item in closure_set:
                next_sym = item.next_symbol()
                if next_sym and next_sym in self.non_terminals:
                    # Compute FIRST(βa) where β is rest of production and a is lookahead
                    beta = item.rhs[item.dot_pos + 1:]
                    
                    # Compute lookaheads for new items
                    lookaheads = set()
                    if beta:
                        # FIRST(β) - for simplicity, take first symbol only
                        for symbol in beta:
                            lookaheads.update(self.first.get(symbol, set()) - {'ε'})
                            # If symbol is terminal or non-nullable non-terminal, stop
                            if symbol in self.terminals:
                                break
                            # Assume non-terminals are not nullable
                            break
                        else:
                            # All of β is nullable (shouldn't happen with our simplification)
                            lookaheads.add(item.lookahead)
                    else:
                        # Just use current lookahead
                        lookaheads.add(item.lookahead)
                    
                    # Add items for all productions of next_sym
                    for production in self.grammar.get(next_sym, []):
                        for la in lookaheads:
                            new_item = LR1Item(next_sym, production, 0, la)
                            if new_item not in closure_set:
                                new_items.add(new_item)
                                changed = True
            
            closure_set.update(new_items)
        
        return closure_set
    
    def goto(self, items, symbol):
        """Compute GOTO(items, symbol) for LR(1)"""
        moved_items = set()
        
        for item in items:
            if item.next_symbol() == symbol:
                moved_items.add(item.advance())
        
        if not moved_items:
            return None
        
        return self.closure(moved_items)
    
    def build_automaton(self):
        """Build LR(1) automaton"""
        print("\n=== BUILDING LR(1) AUTOMATON ===\n")
        print("Note: LR(1) may generate more states than LR(0)/SLR(1)")
        
        # Create initial item with $ as lookahead
        initial_item = LR1Item(self.augmented_start, 
                              self.grammar[self.augmented_start][0], 
                              0,
                              '$')
        initial_items = self.closure({initial_item})
        initial_items_frozen = frozenset(initial_items)
        
        initial_state = LR1State(0, initial_items)
        self.states.append(initial_state)
        self.state_map[initial_items_frozen] = initial_state
        
        queue = deque([initial_state])
        
        while queue:
            current_state = queue.popleft()
            
            # Find all symbols that can be shifted
            symbols_to_process = set()
            for item in current_state.items:
                next_sym = item.next_symbol()
                if next_sym:
                    symbols_to_process.add(next_sym)
            
            for symbol in symbols_to_process:
                goto_items = self.goto(current_state.items, symbol)
                
                if goto_items:
                    goto_items_frozen = frozenset(goto_items)
                    
                    if goto_items_frozen in self.state_map:
                        next_state = self.state_map[goto_items_frozen]
                    else:
                        next_state = LR1State(len(self.states), goto_items)
                        self.states.append(next_state)
                        self.state_map[goto_items_frozen] = next_state
                        queue.append(next_state)
                    
                    self.transitions[(current_state.id, symbol)] = next_state.id
        
        print(f"Created {len(self.states)} states")
        print(f"Created {len(self.transitions)} transitions")
    
    def build_parsing_tables(self):
        """Build CLR (LR(1)) parsing tables"""
        print("\n=== BUILDING CLR (LR(1)) PARSING TABLES ===\n")
        
        for state in self.states:
            state_actions = defaultdict(set)
            
            for item in state.items:
                if item.is_complete():
                    if item.lhs == self.augmented_start:
                        state_actions['$'].add('accept')
                    else:
                        prod_num = self.get_production_number(item.lhs, item.rhs)
                        reduce_action = f"r{prod_num}"
                        
                        # CLR: Only add reduce for the specific lookahead symbol
                        state_actions[item.lookahead].add(reduce_action)
                else:
                    next_sym = item.next_symbol()
                    if next_sym in self.terminals:
                        if (state.id, next_sym) in self.transitions:
                            next_state = self.transitions[(state.id, next_sym)]
                            shift_action = f"s{next_state}"
                            state_actions[next_sym].add(shift_action)
            
            for terminal, actions in state_actions.items():
                actions_list = sorted(list(actions))
                
                if len(actions_list) == 1:
                    self.action_table[state.id][terminal] = actions_list[0]
                else:
                    conflict_type = 'reduce-reduce' if all(a.startswith('r') or a == 'accept' for a in actions_list) else 'shift-reduce'
                    
                    self.conflicts.append({
                        'state': state.id,
                        'symbol': terminal,
                        'type': conflict_type,
                        'action1': actions_list[0],
                        'action2': actions_list[1] if len(actions_list) > 1 else ''
                    })
                    
                    self.action_table[state.id][terminal] = ' / '.join(actions_list)
            
            for non_terminal in self.non_terminals:
                if (state.id, non_terminal) in self.transitions:
                    next_state = self.transitions[(state.id, non_terminal)]
                    self.goto_table[state.id][non_terminal] = next_state
        
        print(f"ACTION table: {len(self.action_table)} states")
        print(f"GOTO table: {len(self.goto_table)} states")
        
        if self.conflicts:
            print(f"\n⚠️  Found {len(self.conflicts)} conflict(s)!")
        else:
            print("\n✓ No conflicts found - Grammar is CLR (LR(1))!")
    
    def get_production_number(self, lhs, rhs):
        """Get production number for A → α
        Production 0 is always the augmented production: ProgramBar → Program
        """
        prod_num = 0
        
        # First, check augmented production (always production 0)
        if lhs == self.augmented_start:
            for prod in self.grammar[self.augmented_start]:
                if prod == rhs:
                    return prod_num
                prod_num += 1
        
        # Then number the rest in order (excluding augmented start)
        for nt in sorted(self.grammar.keys()):
            if nt == self.augmented_start:
                continue
            for prod in self.grammar[nt]:
                if nt == lhs and prod == rhs:
                    return prod_num
                prod_num += 1
        return -1
    
    def is_clr(self):
        """Check if grammar is CLR (LR(1))"""
        return len(self.conflicts) == 0
    
    def generate_excel(self, filename='CLR_minimal/clr_analysis.xlsx'):
        """Generate Excel file with complete analysis"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        self.create_result_sheet(wb)
        self.create_grammar_sheet(wb)
        self.create_first_sheet(wb)
        self.create_states_sheet(wb)
        self.create_action_table_sheet(wb)
        self.create_goto_table_sheet(wb)
        self.create_transitions_sheet(wb)
        
        wb.save(filename)
        print(f"\n✅ Excel file created: {filename}")
    
    def create_result_sheet(self, wb):
        """Create result sheet"""
        ws = wb.create_sheet("Result", 0)
        
        ws['A1'] = "CLR (LR(1)) Parser Analysis Result"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        ws['A3'] = "Is CLR (LR(1))?"
        ws['A3'].font = Font(bold=True, size=14)
        
        is_clr = self.is_clr()
        ws['B3'] = "YES ✓" if is_clr else "NO ✗"
        ws['B3'].font = Font(bold=True, size=14, color="008000" if is_clr else "FF0000")
        
        ws['A5'] = "Number of States:"
        ws['A5'].font = Font(bold=True)
        ws['B5'] = len(self.states)
        
        ws['A6'] = "Number of Conflicts:"
        ws['A6'].font = Font(bold=True)
        ws['B6'] = len(self.conflicts)
        
        if self.conflicts:
            ws['A8'] = "Conflicts Details:"
            ws['A8'].font = Font(bold=True, size=12, color="FF0000")
            ws.merge_cells('A8:E8')
            
            ws['A10'] = "State"
            ws['B10'] = "Symbol"
            ws['C10'] = "Conflict Type"
            ws['D10'] = "Action 1"
            ws['E10'] = "Action 2"
            
            for cell in ['A10', 'B10', 'C10', 'D10', 'E10']:
                ws[cell].font = Font(bold=True, size=11)
                ws[cell].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                ws[cell].alignment = Alignment(horizontal='center', vertical='center')
            
            row = 11
            for conflict in self.conflicts:
                ws[f'A{row}'] = conflict['state']
                ws[f'B{row}'] = conflict['symbol']
                ws[f'C{row}'] = conflict['type']
                ws[f'D{row}'] = conflict['action1']
                ws[f'E{row}'] = conflict['action2']
                
                for col in ['A', 'B', 'C', 'D', 'E']:
                    ws[f'{col}{row}'].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                
                row += 1
        else:
            ws['A8'] = "✓ No conflicts found - Grammar is CLR (LR(1))!"
            ws['A8'].font = Font(bold=True, size=12, color="008000")
            ws.merge_cells('A8:D8')
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
    
    def create_grammar_sheet(self, wb):
        """Create grammar sheet with correct production numbering"""
        ws = wb.create_sheet("Grammar")
        
        ws['A1'] = "Production #"
        ws['B1'] = "Non-Terminal"
        ws['C1'] = "Production"
        
        for cell in ['A1', 'B1', 'C1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
        
        row = 2
        prod_num = 0
        
        # First, add augmented production (production 0)
        for prod in self.grammar[self.augmented_start]:
            ws[f'A{row}'] = prod_num
            ws[f'B{row}'] = self.augmented_start
            ws[f'C{row}'] = f"{self.augmented_start} → {' '.join(prod)}"
            ws[f'A{row}'].fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            ws[f'B{row}'].fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            ws[f'C{row}'].fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            row += 1
            prod_num += 1
        
        # Then add the rest in sorted order
        for nt in sorted(self.grammar.keys()):
            if nt == self.augmented_start:
                continue
            for prod in self.grammar[nt]:
                ws[f'A{row}'] = prod_num
                ws[f'B{row}'] = nt
                ws[f'C{row}'] = f"{nt} → {' '.join(prod)}"
                row += 1
                prod_num += 1
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
    
    def create_first_sheet(self, wb):
        """Create FIRST sets sheet"""
        ws = wb.create_sheet("FIRST Sets")
        
        ws['A1'] = "Symbol"
        ws['B1'] = "FIRST Set"
        ws['C1'] = "Explanation"
        
        for cell in ['A1', 'B1', 'C1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
        
        row = 2
        for A in sorted(self.non_terminals):
            ws[f'A{row}'] = A
            ws[f'B{row}'] = ', '.join(sorted(self.first.get(A, set())))
            ws[f'C{row}'] = "Used for computing lookaheads in CLR"
            row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40
    
    def create_states_sheet(self, wb):
        """Create states sheet"""
        ws = wb.create_sheet("States")
        
        ws['A1'] = "State"
        ws['B1'] = "LR(1) Items"
        
        for cell in ['A1', 'B1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
        
        row = 2
        for state in self.states:
            ws[f'A{row}'] = f"State {state.id}"
            ws[f'A{row}'].font = Font(bold=True)
            
            items_text = "\n".join(str(item) for item in sorted(state.items, key=str))
            ws[f'B{row}'] = items_text
            ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            
            ws.row_dimensions[row].height = 15 * len(state.items)
            
            row += 1
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 60
    
    def create_action_table_sheet(self, wb):
        """Create ACTION table sheet"""
        ws = wb.create_sheet("ACTION Table")
        
        terminals_sorted = sorted(self.terminals)
        
        ws['A1'] = "State"
        for col_idx, terminal in enumerate(terminals_sorted, start=2):
            ws.cell(1, col_idx, terminal)
        
        for col_idx in range(1, len(terminals_sorted) + 2):
            cell = ws.cell(1, col_idx)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for state in self.states:
            row = state.id + 2
            ws.cell(row, 1, state.id)
            ws.cell(row, 1).font = Font(bold=True)
            
            for col_idx, terminal in enumerate(terminals_sorted, start=2):
                if terminal in self.action_table[state.id]:
                    action = self.action_table[state.id][terminal]
                    ws.cell(row, col_idx, action)
                    
                    if '/' in action:
                        ws.cell(row, col_idx).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        ws.cell(row, col_idx).font = Font(color="FFFFFF", bold=True)
                else:
                    ws.cell(row, col_idx, "-")
                    ws.cell(row, col_idx).alignment = Alignment(horizontal='center')
        
        ws.column_dimensions['A'].width = 10
        for col_idx in range(2, len(terminals_sorted) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12
    
    def create_goto_table_sheet(self, wb):
        """Create GOTO table sheet"""
        ws = wb.create_sheet("GOTO Table")
        
        non_terminals_sorted = sorted(self.non_terminals)
        
        ws['A1'] = "State"
        for col_idx, non_terminal in enumerate(non_terminals_sorted, start=2):
            ws.cell(1, col_idx, non_terminal)
        
        for col_idx in range(1, len(non_terminals_sorted) + 2):
            cell = ws.cell(1, col_idx)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for state in self.states:
            row = state.id + 2
            ws.cell(row, 1, state.id)
            ws.cell(row, 1).font = Font(bold=True)
            
            for col_idx, non_terminal in enumerate(non_terminals_sorted, start=2):
                if non_terminal in self.goto_table[state.id]:
                    goto_state = self.goto_table[state.id][non_terminal]
                    ws.cell(row, col_idx, goto_state)
                else:
                    ws.cell(row, col_idx, "-")
                    ws.cell(row, col_idx).alignment = Alignment(horizontal='center')
        
        ws.column_dimensions['A'].width = 10
        for col_idx in range(2, len(non_terminals_sorted) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12
    
    def create_transitions_sheet(self, wb):
        """Create transitions sheet"""
        ws = wb.create_sheet("Transitions")
        
        ws['A1'] = "From State"
        ws['B1'] = "Symbol"
        ws['C1'] = "To State"
        
        for cell in ['A1', 'B1', 'C1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
        
        row = 2
        for (from_state, symbol), to_state in sorted(self.transitions.items()):
            ws[f'A{row}'] = from_state
            ws[f'B{row}'] = symbol
            ws[f'C{row}'] = to_state
            row += 1
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15


def main():
    from grammar import grammar
    
    print("="*70)
    print("CLR (CANONICAL LR / LR(1)) PARSER ANALYZER")
    print("="*70)
    
    analyzer = CLRAnalyzer(grammar)
    
    print(f"\nGrammar has {len(analyzer.non_terminals)} non-terminals and {len(analyzer.terminals)} terminals")
    print(f"Start symbol: {analyzer.start_symbol}")
    print(f"Augmented start: {analyzer.augmented_start}")
    
    analyzer.build_automaton()
    analyzer.build_parsing_tables()
    analyzer.generate_excel()
    
    print("\n" + "="*70)
    print("FINAL RESULT")
    print("="*70)
    
    if analyzer.is_clr():
        print("✅ The grammar IS CLR (LR(1))")
        print("\n🎉 CLR resolved all conflicts!")
        print("This is the most powerful LR parser.")
    else:
        print("❌ The grammar IS NOT CLR (LR(1))")
        print(f"\nFound {len(analyzer.conflicts)} conflict(s):")
        for i, conflict in enumerate(analyzer.conflicts, 1):
            print(f"\n  Conflict {i}:")
            print(f"    State: {conflict['state']}")
            print(f"    Symbol: {conflict['symbol']}")
            print(f"    Type: {conflict['type']}")
            print(f"    Action 1: {conflict['action1']}")
            print(f"    Action 2: {conflict['action2']}")
    
    print("\n" + "="*70)


if __name__ == "__main__":
    main()

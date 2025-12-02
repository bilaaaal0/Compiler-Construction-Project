"""
LALR(1) Parser Analyzer
Merges compatible CLR states - Industry standard (Yacc, Bison)

Note: This implementation builds CLR first then merges states.
In practice, LALR can be built more efficiently.
"""

from collections import defaultdict, deque
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Import CLR analyzer from parent directory
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'CLR_minimal'))
from clr_analyzer import CLRAnalyzer


class LALRAnalyzer:
    def __init__(self, grammar_text):
        # First build CLR automaton
        print("Building CLR automaton first...")
        self.clr = CLRAnalyzer(grammar_text)
        self.clr.build_automaton()
        
        # Copy grammar info
        self.grammar = self.clr.grammar
        self.terminals = self.clr.terminals
        self.non_terminals = self.clr.non_terminals
        self.start_symbol = self.clr.start_symbol
        self.augmented_start = self.clr.augmented_start
        self.first = self.clr.first
        
        # LALR states (merged from CLR)
        self.states = []
        self.state_map = {}
        self.transitions = {}
        self.action_table = defaultdict(dict)
        self.goto_table = defaultdict(dict)
        self.conflicts = []
        
        self.clr_to_lalr_map = {}  # Maps CLR state id to LALR state id

    
    def merge_clr_states(self):
        """Merge CLR states with same core to create LALR states"""
        print("\n=== MERGING CLR STATES TO CREATE LALR ===\n")
        print(f"CLR has {len(self.clr.states)} states")
        
        # Group CLR states by core
        core_groups = defaultdict(list)
        for clr_state in self.clr.states:
            # Core is the set of items without lookaheads
            core = frozenset((item.lhs, tuple(item.rhs), item.dot_pos) 
                           for item in clr_state.items)
            core_groups[core].append(clr_state)
        
        print(f"Found {len(core_groups)} unique cores")
        
        # Create LALR states by merging CLR states with same core
        lalr_id = 0
        for core, clr_states_list in core_groups.items():
            # Merge all items from CLR states with same core
            merged_items = set()
            for clr_state in clr_states_list:
                merged_items.update(clr_state.items)
            
            # Create LALR state
            lalr_state = type('LALRState', (), {
                'id': lalr_id,
                'items': frozenset(merged_items),
                'clr_states': [s.id for s in clr_states_list]
            })()
            
            self.states.append(lalr_state)
            
            # Map each CLR state to this LALR state
            for clr_state in clr_states_list:
                self.clr_to_lalr_map[clr_state.id] = lalr_id
            
            lalr_id += 1
        
        print(f"Created {len(self.states)} LALR states")
        print(f"Reduction: {len(self.clr.states)} → {len(self.states)} states")
        
        # Build LALR transitions from CLR transitions
        for (from_clr, symbol), to_clr in self.clr.transitions.items():
            from_lalr = self.clr_to_lalr_map[from_clr]
            to_lalr = self.clr_to_lalr_map[to_clr]
            self.transitions[(from_lalr, symbol)] = to_lalr
        
        print(f"Created {len(self.transitions)} LALR transitions")
    
    def build_parsing_tables(self):
        """Build LALR(1) parsing tables"""
        print("\n=== BUILDING LALR(1) PARSING TABLES ===\n")
        
        for state in self.states:
            state_actions = defaultdict(set)
            
            for item in state.items:
                if item.is_complete():
                    if item.lhs == self.augmented_start:
                        state_actions['$'].add('accept')
                    else:
                        prod_num = self.get_production_number(item.lhs, item.rhs)
                        reduce_action = f"r{prod_num}"
                        # LALR: reduce for specific lookahead
                        state_actions[item.lookahead].add(reduce_action)
                else:
                    next_sym = item.next_symbol()
                    if next_sym in self.terminals:
                        if (state.id, next_sym) in self.transitions:
                            next_state = self.transitions[(state.id, next_sym)]
                            shift_action = f"s{next_state}"
                            state_actions[next_sym].add(shift_action)
            
            for terminal, actions in state_actions.items():
                actions_list = sorted(list(actions))
                
                if len(actions_list) == 1:
                    self.action_table[state.id][terminal] = actions_list[0]
                else:
                    conflict_type = 'reduce-reduce' if all(a.startswith('r') or a == 'accept' for a in actions_list) else 'shift-reduce'
                    
                    self.conflicts.append({
                        'state': state.id,
                        'symbol': terminal,
                        'type': conflict_type,
                        'action1': actions_list[0],
                        'action2': actions_list[1] if len(actions_list) > 1 else ''
                    })
                    
                    self.action_table[state.id][terminal] = ' / '.join(actions_list)
            
            for non_terminal in self.non_terminals:
                if (state.id, non_terminal) in self.transitions:
                    next_state = self.transitions[(state.id, non_terminal)]
                    self.goto_table[state.id][non_terminal] = next_state
        
        print(f"ACTION table: {len(self.action_table)} states")
        print(f"GOTO table: {len(self.goto_table)} states")
        
        if self.conflicts:
            print(f"\n⚠️  Found {len(self.conflicts)} conflict(s)!")
        else:
            print("\n✓ No conflicts found - Grammar is LALR(1)!")
    
    def get_production_number(self, lhs, rhs):
        """Get production number
        Production 0 is always the augmented production: ProgramBar → Program
        """
        prod_num = 0
        
        # First, check augmented production (always production 0)
        if lhs == self.augmented_start:
            for prod in self.grammar[self.augmented_start]:
                if prod == rhs:
                    return prod_num
                prod_num += 1
        
        # Then number the rest in order (excluding augmented start)
        for nt in sorted(self.grammar.keys()):
            if nt == self.augmented_start:
                continue
            for prod in self.grammar[nt]:
                if nt == lhs and prod == rhs:
                    return prod_num
                prod_num += 1
        return -1
    
    def is_lalr1(self):
        """Check if grammar is LALR(1)"""
        return len(self.conflicts) == 0
    
    def generate_excel(self, filename='LALR_minimal/lalr_analysis.xlsx'):
        """Generate Excel file"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        self.create_result_sheet(wb)
        self.create_comparison_sheet(wb)
        self.create_grammar_sheet(wb)
        self.create_states_sheet(wb)
        self.create_action_table_sheet(wb)
        self.create_goto_table_sheet(wb)
        
        wb.save(filename)
        print(f"\n✅ Excel file created: {filename}")
    
    def create_result_sheet(self, wb):
        """Create result sheet"""
        ws = wb.create_sheet("Result", 0)
        
        ws['A1'] = "LALR(1) Parser Analysis Result"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        ws['A3'] = "Is LALR(1)?"
        ws['A3'].font = Font(bold=True, size=14)
        
        is_lalr1 = self.is_lalr1()
        ws['B3'] = "YES ✓" if is_lalr1 else "NO ✗"
        ws['B3'].font = Font(bold=True, size=14, color="008000" if is_lalr1 else "FF0000")
        
        ws['A5'] = "CLR States:"
        ws['A5'].font = Font(bold=True)
        ws['B5'] = len(self.clr.states)
        
        ws['A6'] = "LALR States:"
        ws['A6'].font = Font(bold=True)
        ws['B6'] = len(self.states)
        
        ws['A7'] = "States Merged:"
        ws['A7'].font = Font(bold=True)
        ws['B7'] = len(self.clr.states) - len(self.states)
        ws['B7'].font = Font(color="008000")
        
        ws['A9'] = "Number of Conflicts:"
        ws['A9'].font = Font(bold=True)
        ws['B9'] = len(self.conflicts)
        
        if self.conflicts:
            ws['A11'] = "Conflicts Details:"
            ws['A11'].font = Font(bold=True, size=12, color="FF0000")
            ws.merge_cells('A11:E11')
            
            ws['A13'] = "State"
            ws['B13'] = "Symbol"
            ws['C13'] = "Type"
            ws['D13'] = "Action 1"
            ws['E13'] = "Action 2"
            
            for cell in ['A13', 'B13', 'C13', 'D13', 'E13']:
                ws[cell].font = Font(bold=True, size=11)
                ws[cell].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                ws[cell].alignment = Alignment(horizontal='center')
            
            row = 14
            for conflict in self.conflicts:
                ws[f'A{row}'] = conflict['state']
                ws[f'B{row}'] = conflict['symbol']
                ws[f'C{row}'] = conflict['type']
                ws[f'D{row}'] = conflict['action1']
                ws[f'E{row}'] = conflict['action2']
                
                for col in ['A', 'B', 'C', 'D', 'E']:
                    ws[f'{col}{row}'].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                row += 1
        else:
            ws['A11'] = "✓ No conflicts - Grammar is LALR(1)!"
            ws['A11'].font = Font(bold=True, size=12, color="008000")
            ws.merge_cells('A11:D11')
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
    
    def create_comparison_sheet(self, wb):
        """Create comparison sheet"""
        ws = wb.create_sheet("CLR vs LALR")
        
        ws['A1'] = "Comparison: CLR vs LALR"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:C1')
        
        ws['A3'] = "Metric"
        ws['B3'] = "CLR"
        ws['C3'] = "LALR"
        
        for cell in ['A3', 'B3', 'C3']:
            ws[cell].font = Font(bold=True)
            ws[cell].fill = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
            ws[cell].font = Font(bold=True, color="FFFFFF")
        
        row = 4
        ws[f'A{row}'] = "States"
        ws[f'B{row}'] = len(self.clr.states)
        ws[f'C{row}'] = len(self.states)
        row += 1
        
        ws[f'A{row}'] = "Conflicts"
        ws[f'B{row}'] = len(self.clr.conflicts)
        ws[f'C{row}'] = len(self.conflicts)
        row += 1
        
        ws[f'A{row}'] = "Accepted?"
        ws[f'B{row}'] = "NO" if len(self.clr.conflicts) > 0 else "YES"
        ws[f'C{row}'] = "NO" if len(self.conflicts) > 0 else "YES"
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
    
    def create_grammar_sheet(self, wb):
        """Create grammar sheet with correct production numbering"""
        ws = wb.create_sheet("Grammar")
        
        ws['A1'] = "Production #"
        ws['B1'] = "Non-Terminal"
        ws['C1'] = "Production"
        
        for cell in ['A1', 'B1', 'C1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
        
        row = 2
        prod_num = 0
        
        # First, add augmented production (production 0)
        for prod in self.grammar[self.augmented_start]:
            ws[f'A{row}'] = prod_num
            ws[f'B{row}'] = self.augmented_start
            ws[f'C{row}'] = f"{self.augmented_start} → {' '.join(prod)}"
            ws[f'A{row}'].fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            ws[f'B{row}'].fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            ws[f'C{row}'].fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            row += 1
            prod_num += 1
        
        # Then add the rest in sorted order
        for nt in sorted(self.grammar.keys()):
            if nt == self.augmented_start:
                continue
            for prod in self.grammar[nt]:
                ws[f'A{row}'] = prod_num
                ws[f'B{row}'] = nt
                ws[f'C{row}'] = f"{nt} → {' '.join(prod)}"
                row += 1
                prod_num += 1
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
    
    def create_states_sheet(self, wb):
        """Create states sheet"""
        ws = wb.create_sheet("States")
        
        ws['A1'] = "State"
        ws['B1'] = "LR(1) Items"
        ws['C1'] = "Merged from CLR States"
        
        for cell in ['A1', 'B1', 'C1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
        
        row = 2
        for state in self.states:
            ws[f'A{row}'] = f"State {state.id}"
            ws[f'A{row}'].font = Font(bold=True)
            
            items_text = "\n".join(str(item) for item in sorted(state.items, key=str))
            ws[f'B{row}'] = items_text
            ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            
            ws[f'C{row}'] = ', '.join(map(str, state.clr_states))
            
            ws.row_dimensions[row].height = 15 * min(len(state.items), 10)
            row += 1
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 20
    
    def create_action_table_sheet(self, wb):
        """Create ACTION table"""
        ws = wb.create_sheet("ACTION Table")
        
        terminals_sorted = sorted(self.terminals)
        
        ws['A1'] = "State"
        for col_idx, terminal in enumerate(terminals_sorted, start=2):
            ws.cell(1, col_idx, terminal)
        
        for col_idx in range(1, len(terminals_sorted) + 2):
            cell = ws.cell(1, col_idx)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
        
        for state in self.states:
            row = state.id + 2
            ws.cell(row, 1, state.id)
            ws.cell(row, 1).font = Font(bold=True)
            
            for col_idx, terminal in enumerate(terminals_sorted, start=2):
                if terminal in self.action_table[state.id]:
                    action = self.action_table[state.id][terminal]
                    ws.cell(row, col_idx, action)
                    
                    if '/' in action:
                        ws.cell(row, col_idx).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        ws.cell(row, col_idx).font = Font(color="FFFFFF", bold=True)
                else:
                    ws.cell(row, col_idx, "-")
                    ws.cell(row, col_idx).alignment = Alignment(horizontal='center')
        
        ws.column_dimensions['A'].width = 10
        for col_idx in range(2, len(terminals_sorted) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12
    
    def create_goto_table_sheet(self, wb):
        """Create GOTO table"""
        ws = wb.create_sheet("GOTO Table")
        
        non_terminals_sorted = sorted(self.non_terminals)
        
        ws['A1'] = "State"
        for col_idx, non_terminal in enumerate(non_terminals_sorted, start=2):
            ws.cell(1, col_idx, non_terminal)
        
        for col_idx in range(1, len(non_terminals_sorted) + 2):
            cell = ws.cell(1, col_idx)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
        
        for state in self.states:
            row = state.id + 2
            ws.cell(row, 1, state.id)
            ws.cell(row, 1).font = Font(bold=True)
            
            for col_idx, non_terminal in enumerate(non_terminals_sorted, start=2):
                if non_terminal in self.goto_table[state.id]:
                    goto_state = self.goto_table[state.id][non_terminal]
                    ws.cell(row, col_idx, goto_state)
                else:
                    ws.cell(row, col_idx, "-")
                    ws.cell(row, col_idx).alignment = Alignment(horizontal='center')
        
        ws.column_dimensions['A'].width = 10
        for col_idx in range(2, len(non_terminals_sorted) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12


def main():
    from grammar import grammar
    
    print("="*70)
    print("LALR(1) PARSER ANALYZER")
    print("Industry Standard - Used in Yacc, Bison")
    print("="*70)
    
    analyzer = LALRAnalyzer(grammar)
    
    print(f"\nGrammar has {len(analyzer.non_terminals)} non-terminals and {len(analyzer.terminals)} terminals")
    
    analyzer.merge_clr_states()
    analyzer.build_parsing_tables()
    analyzer.generate_excel()
    
    print("\n" + "="*70)
    print("FINAL RESULT")
    print("="*70)
    
    if analyzer.is_lalr1():
        print("✅ The grammar IS LALR(1)")
        print("\n🎉 LALR(1) resolved all conflicts!")
        print("This is the industry standard parser.")
    else:
        print("❌ The grammar IS NOT LALR(1)")
        print(f"\nFound {len(analyzer.conflicts)} conflict(s):")
        for i, conflict in enumerate(analyzer.conflicts, 1):
            print(f"\n  Conflict {i}:")
            print(f"    State: {conflict['state']}")
            print(f"    Symbol: {conflict['symbol']}")
            print(f"    Type: {conflict['type']}")
            print(f"    Action 1: {conflict['action1']}")
            print(f"    Action 2: {conflict['action2']}")
    
    print(f"\nState Reduction: {len(analyzer.clr.states)} (CLR) → {len(analyzer.states)} (LALR)")
    print(f"Savings: {len(analyzer.clr.states) - len(analyzer.states)} states ({100 * (len(analyzer.clr.states) - len(analyzer.states)) / len(analyzer.clr.states):.1f}%)")
    
    print("\n" + "="*70)


if __name__ == "__main__":
    main()

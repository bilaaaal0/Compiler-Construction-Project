"""
LL(1) Grammar Analyzer for Minimal Grammar
Simplified version for hand-written parsing tables
"""

from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class LL1Analyzer:
    def __init__(self, grammar_text):
        self.original_grammar = {}
        self.grammar = {}
        self.terminals = set()
        self.non_terminals = set()
        self.nullable = set()
        self.first = defaultdict(set)
        self.follow = defaultdict(set)
        self.parsing_table = defaultdict(dict)
        self.conflicts = []
        self.parse_grammar(grammar_text)
        
    def parse_grammar(self, grammar_text):
        """Parse the grammar from text format"""
        lines = grammar_text.strip().split('\n')
        current_lhs = None
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
                
            if '→' in line:
                parts = line.split('→')
                lhs = parts[0].strip()
                rhs = parts[1].strip()
                current_lhs = lhs
                self.non_terminals.add(lhs)
                
                if lhs not in self.original_grammar:
                    self.original_grammar[lhs] = []
                self.original_grammar[lhs].append(self.parse_production(rhs))
                
            elif '|' in line:
                rhs = line.split('|')[1].strip()
                if current_lhs:
                    self.original_grammar[current_lhs].append(self.parse_production(rhs))
        
        # Extract terminals
        for lhs, productions in self.original_grammar.items():
            for prod in productions:
                for symbol in prod:
                    if symbol not in self.non_terminals and symbol != 'ε':
                        self.terminals.add(symbol)
        
        # Add end marker
        self.terminals.add('$')
        
        # Copy to working grammar
        self.grammar = {k: [p[:] for p in v] for k, v in self.original_grammar.items()}
    
    def parse_production(self, rhs):
        """Parse a production right-hand side into symbols"""
        if rhs == 'ε':
            return ['ε']
        
        symbols = []
        i = 0
        while i < len(rhs):
            if rhs[i] == "'":
                # Terminal in quotes
                j = i + 1
                while j < len(rhs) and rhs[j] != "'":
                    j += 1
                symbols.append(rhs[i:j+1])
                i = j + 1
            elif rhs[i].isspace():
                i += 1
            else:
                # Non-terminal or identifier
                j = i
                while j < len(rhs) and not rhs[j].isspace() and rhs[j] != "'":
                    j += 1
                symbol = rhs[i:j]
                if symbol:
                    symbols.append(symbol)
                i = j
        
        return symbols
    
    def compute_nullable(self):
        """Compute NULLABLE set"""
        print("\n=== COMPUTING NULLABLE ===\n")
        
        changed = True
        while changed:
            changed = False
            for A in self.non_terminals:
                if A in self.nullable:
                    continue
                
                for production in self.grammar.get(A, []):
                    if production == ['ε']:
                        self.nullable.add(A)
                        changed = True
                        break
                    
                    if all(symbol in self.nullable for symbol in production if symbol in self.non_terminals):
                        if all(symbol in self.non_terminals for symbol in production):
                            self.nullable.add(A)
                            changed = True
                            break
        
        print(f"NULLABLE = {{{', '.join(sorted(self.nullable))}}}")
    
    def compute_first(self):
        """Compute FIRST sets"""
        print("\n=== COMPUTING FIRST SETS ===\n")
        
        # Initialize FIRST for terminals
        for t in self.terminals:
            self.first[t].add(t)
        
        # Compute FIRST for non-terminals
        changed = True
        while changed:
            changed = False
            for A in self.non_terminals:
                for production in self.grammar.get(A, []):
                    if production == ['ε']:
                        if 'ε' not in self.first[A]:
                            self.first[A].add('ε')
                            changed = True
                    else:
                        for symbol in production:
                            # Add FIRST(symbol) - {ε} to FIRST(A)
                            before = len(self.first[A])
                            self.first[A].update(self.first[symbol] - {'ε'})
                            if len(self.first[A]) > before:
                                changed = True
                            
                            # If symbol is not nullable, stop
                            if symbol not in self.nullable:
                                break
                        else:
                            # All symbols are nullable
                            if 'ε' not in self.first[A]:
                                self.first[A].add('ε')
                                changed = True
        
        for A in sorted(self.non_terminals):
            print(f"FIRST({A}) = {{{', '.join(sorted(self.first[A]))}}}")
    
    def compute_follow(self):
        """Compute FOLLOW sets"""
        print("\n=== COMPUTING FOLLOW SETS ===\n")
        
        # Start symbol gets $
        start_symbol = list(self.grammar.keys())[0]
        self.follow[start_symbol].add('$')
        
        changed = True
        iterations = 0
        max_iterations = 100  # Safety limit
        
        while changed and iterations < max_iterations:
            changed = False
            iterations += 1
            
            for A in self.non_terminals:
                for production in self.grammar.get(A, []):
                    # Skip epsilon productions
                    if production == ['ε']:
                        continue
                    
                    for i, symbol in enumerate(production):
                        # Only process non-terminals
                        if symbol not in self.non_terminals:
                            continue
                        
                        # Get what comes after this symbol (β)
                        beta = production[i+1:]
                        
                        if beta:
                            # Case 1: A → αBβ where β is not empty
                            # Add FIRST(β) - {ε} to FOLLOW(B)
                            all_nullable = True
                            for b_symbol in beta:
                                before_size = len(self.follow[symbol])
                                self.follow[symbol].update(self.first.get(b_symbol, set()) - {'ε'})
                                if len(self.follow[symbol]) > before_size:
                                    changed = True
                                
                                # Check if this symbol is nullable
                                if b_symbol not in self.nullable:
                                    all_nullable = False
                                    break
                            
                            # Case 2: If all of β is nullable, add FOLLOW(A) to FOLLOW(B)
                            if all_nullable and A != symbol:  # Avoid self-reference in same iteration
                                before_size = len(self.follow[symbol])
                                self.follow[symbol].update(self.follow[A])
                                if len(self.follow[symbol]) > before_size:
                                    changed = True
                        else:
                            # Case 3: A → αB (B is last)
                            # Add FOLLOW(A) to FOLLOW(B)
                            if A != symbol:  # Avoid self-reference in same iteration
                                before_size = len(self.follow[symbol])
                                self.follow[symbol].update(self.follow[A])
                                if len(self.follow[symbol]) > before_size:
                                    changed = True
        
        if iterations >= max_iterations:
            print(f"⚠️  Warning: FOLLOW computation did not converge after {max_iterations} iterations")
        else:
            print(f"FOLLOW computation converged after {iterations} iterations\n")
        
        for A in sorted(self.non_terminals):
            follow_set = self.follow.get(A, set())
            if follow_set:
                print(f"FOLLOW({A}) = {{{', '.join(sorted(follow_set))}}}")
            else:
                print(f"FOLLOW({A}) = {{}} (empty!)")
    
    def compute_first_of_string(self, symbols):
        """Compute FIRST of a string of symbols"""
        result = set()
        
        for symbol in symbols:
            result.update(self.first[symbol] - {'ε'})
            if symbol not in self.nullable:
                break
        else:
            result.add('ε')
        
        return result
    
    def build_parsing_table(self):
        """Build LL(1) predictive parsing table"""
        print("\n=== BUILDING PARSING TABLE ===\n")
        
        for A in self.non_terminals:
            for production in self.grammar.get(A, []):
                # Compute FIRST(α)
                first_alpha = self.compute_first_of_string(production)
                
                # For each terminal in FIRST(α), add A → α to table
                for terminal in first_alpha - {'ε'}:
                    if terminal in self.parsing_table[A]:
                        # Conflict!
                        self.conflicts.append({
                            'non_terminal': A,
                            'terminal': terminal,
                            'production1': self.parsing_table[A][terminal],
                            'production2': production
                        })
                        # Use format: S→A / S→B
                        existing = ' '.join(self.parsing_table[A][terminal])
                        new = ' '.join(production)
                        self.parsing_table[A][terminal] = [f"{A}→{existing} / {A}→{new}"]
                    else:
                        self.parsing_table[A][terminal] = production
                
                # If ε in FIRST(α), add A → α for each terminal in FOLLOW(A)
                if 'ε' in first_alpha:
                    for terminal in self.follow[A]:
                        if terminal in self.parsing_table[A]:
                            # Conflict!
                            self.conflicts.append({
                                'non_terminal': A,
                                'terminal': terminal,
                                'production1': self.parsing_table[A][terminal],
                                'production2': production
                            })
                            existing = ' '.join(self.parsing_table[A][terminal])
                            new = ' '.join(production)
                            self.parsing_table[A][terminal] = [f"{A}→{existing} / {A}→{new}"]
                        else:
                            self.parsing_table[A][terminal] = production
        
        print("Parsing table built successfully.")
        if self.conflicts:
            print(f"\n⚠️  Found {len(self.conflicts)} conflict(s)!")
    
    def is_ll1(self):
        """Check if grammar is LL(1)"""
        return len(self.conflicts) == 0
    
    def generate_excel(self, filename='LL1_minimal/grammar_analysis.xlsx'):
        """Generate Excel file with complete analysis"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        self.create_result_sheet(wb)
        self.create_grammar_sheet(wb)
        self.create_nullable_sheet(wb)
        self.create_first_sheet(wb)
        self.create_follow_sheet(wb)
        self.create_parsing_table_sheet(wb)
        self.create_hand_written_guide_sheet(wb)
        
        wb.save(filename)
        print(f"\n✅ Excel file created: {filename}")
    
    def create_result_sheet(self, wb):
        """Create result sheet"""
        ws = wb.create_sheet("Result", 0)
        
        ws['A1'] = "LL(1) Grammar Analysis Result"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        ws['A3'] = "Is LL(1)?"
        ws['A3'].font = Font(bold=True, size=14)
        
        is_ll1 = self.is_ll1()
        ws['B3'] = "YES ✓" if is_ll1 else "NO ✗"
        ws['B3'].font = Font(bold=True, size=14, color="008000" if is_ll1 else "FF0000")
        
        ws['A5'] = "Number of Conflicts:"
        ws['A5'].font = Font(bold=True)
        ws['B5'] = len(self.conflicts)
        
        if self.conflicts:
            ws['A7'] = "Conflicts Details:"
            ws['A7'].font = Font(bold=True, size=12, color="FF0000")
            ws.merge_cells('A7:D7')
            
            ws['A9'] = "Non-Terminal"
            ws['B9'] = "Terminal"
            ws['C9'] = "Production 1"
            ws['D9'] = "Production 2"
            
            for cell in ['A9', 'B9', 'C9', 'D9']:
                ws[cell].font = Font(bold=True, size=11)
                ws[cell].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                ws[cell].alignment = Alignment(horizontal='center', vertical='center')
            
            row = 10
            for conflict in self.conflicts:
                ws[f'A{row}'] = conflict['non_terminal']
                ws[f'B{row}'] = conflict['terminal']
                
                prod1 = ' '.join(conflict['production1'])
                prod2 = ' '.join(conflict['production2'])
                
                ws[f'C{row}'] = f"{conflict['non_terminal']} → {prod1}"
                ws[f'D{row}'] = f"{conflict['non_terminal']} → {prod2}"
                
                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{row}'].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                
                row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 40
    
    def create_grammar_sheet(self, wb):
        """Create grammar sheet"""
        ws = wb.create_sheet("Grammar")
        
        ws['A1'] = "Non-Terminal"
        ws['B1'] = "Productions"
        
        for cell in ['A1', 'B1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
        
        row = 2
        for A in self.grammar.keys():
            ws[f'A{row}'] = A
            productions = ' | '.join([' '.join(p) for p in self.grammar[A]])
            ws[f'B{row}'] = productions
            row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 60
    
    def create_nullable_sheet(self, wb):
        """Create NULLABLE sheet"""
        ws = wb.create_sheet("NULLABLE")
        
        ws['A1'] = "NULLABLE Non-Terminals"
        ws['A1'].font = Font(bold=True, size=12)
        ws['A1'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        row = 2
        for A in sorted(self.nullable):
            ws[f'A{row}'] = A
            row += 1
        
        if not self.nullable:
            ws['A2'] = "(none)"
        
        ws.column_dimensions['A'].width = 25
    
    def create_first_sheet(self, wb):
        """Create FIRST sets sheet"""
        ws = wb.create_sheet("FIRST Sets")
        
        ws['A1'] = "Non-Terminal"
        ws['B1'] = "FIRST Set"
        
        for cell in ['A1', 'B1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
        
        row = 2
        for A in sorted(self.non_terminals):
            ws[f'A{row}'] = A
            ws[f'B{row}'] = ', '.join(sorted(self.first[A]))
            row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
    
    def create_follow_sheet(self, wb):
        """Create FOLLOW sets sheet"""
        ws = wb.create_sheet("FOLLOW Sets")
        
        ws['A1'] = "Non-Terminal"
        ws['B1'] = "FOLLOW Set"
        
        for cell in ['A1', 'B1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
        
        row = 2
        for A in sorted(self.non_terminals):
            ws[f'A{row}'] = A
            ws[f'B{row}'] = ', '.join(sorted(self.follow[A]))
            row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
    
    def create_parsing_table_sheet(self, wb):
        """Create parsing table sheet"""
        ws = wb.create_sheet("Parsing Table")
        
        terminals_sorted = sorted(self.terminals)
        non_terminals_sorted = sorted(self.non_terminals)
        
        ws['A1'] = "Non-Terminal"
        for col_idx, terminal in enumerate(terminals_sorted, start=2):
            ws.cell(1, col_idx, terminal)
        
        for col_idx in range(1, len(terminals_sorted) + 2):
            cell = ws.cell(1, col_idx)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row_idx, non_terminal in enumerate(non_terminals_sorted, start=2):
            ws.cell(row_idx, 1, non_terminal)
            ws.cell(row_idx, 1).font = Font(bold=True)
            
            for col_idx, terminal in enumerate(terminals_sorted, start=2):
                if terminal in self.parsing_table[non_terminal]:
                    production = self.parsing_table[non_terminal][terminal]
                    prod_str = ' '.join(production)
                    
                    if '/' in prod_str:
                        ws.cell(row_idx, col_idx, prod_str)
                        ws.cell(row_idx, col_idx).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        ws.cell(row_idx, col_idx).font = Font(color="FFFFFF", bold=True)
                    else:
                        ws.cell(row_idx, col_idx, f"{non_terminal} → {prod_str}")
                else:
                    ws.cell(row_idx, col_idx, "-")
                    ws.cell(row_idx, col_idx).alignment = Alignment(horizontal='center', vertical='center')
        
        ws.column_dimensions['A'].width = 20
        for col_idx in range(2, len(terminals_sorted) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 25
    
    def create_hand_written_guide_sheet(self, wb):
        """Create guide for hand-written parsing table"""
        ws = wb.create_sheet("Hand-Written Guide")
        
        ws['A1'] = "Step-by-Step Guide for Hand-Written Parsing Table"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        row = 3
        steps = [
            ("Step 1: Write the Grammar", "Copy the grammar productions clearly"),
            ("Step 2: Compute NULLABLE", "Find which non-terminals can derive ε"),
            ("Step 3: Compute FIRST Sets", "For each non-terminal, find what terminals can start it"),
            ("Step 4: Compute FOLLOW Sets", "For each non-terminal, find what terminals can follow it"),
            ("Step 5: Build Parsing Table", "Use FIRST and FOLLOW to fill the table"),
            ("Step 6: Check for Conflicts", "Look for cells with multiple entries"),
        ]
        
        for step, description in steps:
            ws[f'A{row}'] = step
            ws[f'A{row}'].font = Font(bold=True, size=11)
            ws[f'B{row}'] = description
            row += 1
        
        row += 2
        ws[f'A{row}'] = "Table Dimensions:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = f"Rows (Non-Terminals): {len(self.non_terminals)}"
        row += 1
        ws[f'A{row}'] = f"Columns (Terminals): {len(self.terminals)}"
        row += 1
        ws[f'A{row}'] = f"Total Cells: {len(self.non_terminals)} × {len(self.terminals)} = {len(self.non_terminals) * len(self.terminals)}"
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50


def main():
    from grammar import grammar
    
    print("="*70)
    print("LL(1) GRAMMAR ANALYZER - MINIMAL VERSION")
    print("="*70)
    
    analyzer = LL1Analyzer(grammar)
    
    print(f"\nGrammar has {len(analyzer.non_terminals)} non-terminals and {len(analyzer.terminals)} terminals")
    
    analyzer.compute_nullable()
    analyzer.compute_first()
    analyzer.compute_follow()
    analyzer.build_parsing_table()
    analyzer.generate_excel()
    
    print("\n" + "="*70)
    print("FINAL RESULT")
    print("="*70)
    
    if analyzer.is_ll1():
        print("✅ The grammar IS LL(1)")
    else:
        print("❌ The grammar IS NOT LL(1)")
        print(f"\nFound {len(analyzer.conflicts)} conflict(s):")
        for i, conflict in enumerate(analyzer.conflicts, 1):
            print(f"\n  Conflict {i}:")
            print(f"    Non-terminal: {conflict['non_terminal']}")
            print(f"    Terminal: {conflict['terminal']}")
            print(f"    Production 1: {' '.join(conflict['production1'])}")
            print(f"    Production 2: {' '.join(conflict['production2'])}")
    
    print("\n" + "="*70)


if __name__ == "__main__":
    main()

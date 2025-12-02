"""
LR(0) Parser Analyzer for Minimal Grammar
Simplified version for hand-written parsing tables
"""

from collections import defaultdict, deque
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class LR0Item:
    """Represents an LR(0) item: A → α·β"""
    def __init__(self, lhs, rhs, dot_pos):
        self.lhs = lhs
        self.rhs = rhs
        self.dot_pos = dot_pos
    
    def __eq__(self, other):
        return (self.lhs == other.lhs and 
                self.rhs == other.rhs and 
                self.dot_pos == other.dot_pos)
    
    def __hash__(self):
        return hash((self.lhs, tuple(self.rhs), self.dot_pos))
    
    def __repr__(self):
        rhs_with_dot = self.rhs[:self.dot_pos] + ['·'] + self.rhs[self.dot_pos:]
        return f"{self.lhs} → {' '.join(rhs_with_dot)}"
    
    def is_complete(self):
        return self.dot_pos >= len(self.rhs)
    
    def next_symbol(self):
        if self.is_complete():
            return None
        return self.rhs[self.dot_pos]
    
    def advance(self):
        return LR0Item(self.lhs, self.rhs, self.dot_pos + 1)


class LR0State:
    """Represents a state in LR(0) automaton"""
    def __init__(self, state_id, items):
        self.id = state_id
        self.items = frozenset(items)
    
    def __eq__(self, other):
        return self.items == other.items
    
    def __hash__(self):
        return hash(self.items)
    
    def __repr__(self):
        return f"State {self.id}:\n" + "\n".join(f"  {item}" for item in self.items)


class LR0Analyzer:
    def __init__(self, grammar_text):
        self.grammar = {}
        self.terminals = set()
        self.non_terminals = set()
        self.start_symbol = None
        self.augmented_start = None
        
        self.states = []
        self.state_map = {}
        self.transitions = {}
        
        self.action_table = defaultdict(dict)
        self.goto_table = defaultdict(dict)
        self.conflicts = []
        
        self.parse_grammar(grammar_text)
        self.augment_grammar()
        
    def parse_grammar(self, grammar_text):
        """Parse the grammar from text format"""
        lines = grammar_text.strip().split('\n')
        current_lhs = None
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
                
            if '→' in line:
                parts = line.split('→')
                lhs = parts[0].strip()
                rhs = parts[1].strip()
                current_lhs = lhs
                self.non_terminals.add(lhs)
                
                if self.start_symbol is None:
                    self.start_symbol = lhs
                
                if lhs not in self.grammar:
                    self.grammar[lhs] = []
                self.grammar[lhs].append(self.parse_production(rhs))
                
            elif '|' in line:
                rhs = line.split('|')[1].strip()
                if current_lhs:
                    self.grammar[current_lhs].append(self.parse_production(rhs))
        
        # Extract terminals
        for lhs, productions in self.grammar.items():
            for prod in productions:
                for symbol in prod:
                    if symbol not in self.non_terminals:
                        self.terminals.add(symbol)
        
        self.terminals.add('$')
    
    def parse_production(self, rhs):
        """Parse a production right-hand side into symbols"""
        symbols = []
        i = 0
        while i < len(rhs):
            if rhs[i] == "'":
                j = i + 1
                while j < len(rhs) and rhs[j] != "'":
                    j += 1
                symbols.append(rhs[i:j+1])
                i = j + 1
            elif rhs[i].isspace():
                i += 1
            else:
                j = i
                while j < len(rhs) and not rhs[j].isspace() and rhs[j] != "'":
                    j += 1
                symbol = rhs[i:j]
                if symbol:
                    symbols.append(symbol)
                i = j
        
        return symbols
    
    def augment_grammar(self):
        """Augment grammar with SBar → S"""
        self.augmented_start = self.start_symbol + "Bar"
        self.grammar[self.augmented_start] = [[self.start_symbol]]
        self.non_terminals.add(self.augmented_start)
    
    def closure(self, items):
        """Compute closure of a set of items"""
        closure_set = set(items)
        changed = True
        
        while changed:
            changed = False
            new_items = set()
            
            for item in closure_set:
                next_sym = item.next_symbol()
                if next_sym and next_sym in self.non_terminals:
                    for production in self.grammar.get(next_sym, []):
                        new_item = LR0Item(next_sym, production, 0)
                        if new_item not in closure_set:
                            new_items.add(new_item)
                            changed = True
            
            closure_set.update(new_items)
        
        return closure_set
    
    def goto(self, items, symbol):
        """Compute GOTO(items, symbol)"""
        moved_items = set()
        
        for item in items:
            if item.next_symbol() == symbol:
                moved_items.add(item.advance())
        
        if not moved_items:
            return None
        
        return self.closure(moved_items)
    
    def build_automaton(self):
        """Build LR(0) automaton"""
        print("\n=== BUILDING LR(0) AUTOMATON ===\n")
        
        initial_item = LR0Item(self.augmented_start, 
                              self.grammar[self.augmented_start][0], 
                              0)
        initial_items = self.closure({initial_item})
        initial_items_frozen = frozenset(initial_items)
        
        initial_state = LR0State(0, initial_items)
        self.states.append(initial_state)
        self.state_map[initial_items_frozen] = initial_state
        
        queue = deque([initial_state])
        
        while queue:
            current_state = queue.popleft()
            
            symbols_to_process = set()
            for item in current_state.items:
                next_sym = item.next_symbol()
                if next_sym:
                    symbols_to_process.add(next_sym)
            
            for symbol in symbols_to_process:
                goto_items = self.goto(current_state.items, symbol)
                
                if goto_items:
                    goto_items_frozen = frozenset(goto_items)
                    
                    if goto_items_frozen in self.state_map:
                        next_state = self.state_map[goto_items_frozen]
                    else:
                        next_state = LR0State(len(self.states), goto_items)
                        self.states.append(next_state)
                        self.state_map[goto_items_frozen] = next_state
                        queue.append(next_state)
                    
                    self.transitions[(current_state.id, symbol)] = next_state.id
        
        print(f"Created {len(self.states)} states")
        print(f"Created {len(self.transitions)} transitions")
    
    def build_parsing_tables(self):
        """Build ACTION and GOTO tables"""
        print("\n=== BUILDING PARSING TABLES ===\n")
        
        for state in self.states:
            state_actions = defaultdict(set)
            
            for item in state.items:
                if item.is_complete():
                    if item.lhs == self.augmented_start:
                        state_actions['$'].add('accept')
                    else:
                        prod_num = self.get_production_number(item.lhs, item.rhs)
                        reduce_action = f"r{prod_num}"
                        for terminal in self.terminals:
                            state_actions[terminal].add(reduce_action)
                else:
                    next_sym = item.next_symbol()
                    if next_sym in self.terminals:
                        if (state.id, next_sym) in self.transitions:
                            next_state = self.transitions[(state.id, next_sym)]
                            shift_action = f"s{next_state}"
                            state_actions[next_sym].add(shift_action)
            
            for terminal, actions in state_actions.items():
                actions_list = sorted(list(actions))
                
                if len(actions_list) == 1:
                    self.action_table[state.id][terminal] = actions_list[0]
                else:
                    conflict_type = 'reduce-reduce' if all(a.startswith('r') or a == 'accept' for a in actions_list) else 'shift-reduce'
                    
                    self.conflicts.append({
                        'state': state.id,
                        'symbol': terminal,
                        'type': conflict_type,
                        'action1': actions_list[0],
                        'action2': actions_list[1] if len(actions_list) > 1 else ''
                    })
                    
                    self.action_table[state.id][terminal] = ' / '.join(actions_list)
            
            for non_terminal in self.non_terminals:
                if (state.id, non_terminal) in self.transitions:
                    next_state = self.transitions[(state.id, non_terminal)]
                    self.goto_table[state.id][non_terminal] = next_state
        
        print(f"ACTION table: {len(self.action_table)} states")
        print(f"GOTO table: {len(self.goto_table)} states")
        
        if self.conflicts:
            print(f"\n⚠️  Found {len(self.conflicts)} conflict(s)!")
        else:
            print("\n✓ No conflicts found!")
    
    def get_production_number(self, lhs, rhs):
        """Get production number for A → α
        Production 0 is always the augmented production: ProgramBar → Program
        """
        prod_num = 0
        
        # First, check augmented production (always production 0)
        if lhs == self.augmented_start:
            for prod in self.grammar[self.augmented_start]:
                if prod == rhs:
                    return prod_num
                prod_num += 1
        
        # Then number the rest in order (excluding augmented start)
        for nt in sorted(self.grammar.keys()):
            if nt == self.augmented_start:
                continue
            for prod in self.grammar[nt]:
                if nt == lhs and prod == rhs:
                    return prod_num
                prod_num += 1
        return -1
    
    def is_lr0(self):
        """Check if grammar is LR(0)"""
        return len(self.conflicts) == 0
    
    def compute_follow(self):
        """Compute FOLLOW sets for SLR(1) analysis"""
        print("\n=== COMPUTING FOLLOW SETS FOR SLR(1) ===\n")
        
        follow = defaultdict(set)
        
        # Start symbol gets $
        follow[self.augmented_start].add('$')
        
        changed = True
        while changed:
            changed = False
            for A in self.non_terminals:
                for production in self.grammar.get(A, []):
                    for i, symbol in enumerate(production):
                        if symbol in self.non_terminals:
                            # Add FIRST(β) - {ε} to FOLLOW(symbol)
                            beta = production[i+1:]
                            
                            if beta:
                                # Compute FIRST of beta
                                for b_symbol in beta:
                                    if b_symbol in self.terminals:
                                        before = len(follow[symbol])
                                        follow[symbol].add(b_symbol)
                                        if len(follow[symbol]) > before:
                                            changed = True
                                        break
                                    elif b_symbol in self.non_terminals:
                                        # For simplicity, assume non-terminals are not nullable
                                        # In full implementation, would check nullable
                                        break
                                else:
                                    # All of β is nullable or empty
                                    before = len(follow[symbol])
                                    follow[symbol].update(follow[A])
                                    if len(follow[symbol]) > before:
                                        changed = True
                            else:
                                # symbol is last
                                before = len(follow[symbol])
                                follow[symbol].update(follow[A])
                                if len(follow[symbol]) > before:
                                    changed = True
        
        for A in sorted(self.non_terminals):
            if A in follow:
                print(f"FOLLOW({A}) = {{{', '.join(sorted(follow[A]))}}}")
        
        return follow
    
    def build_slr_parsing_tables(self):
        """Build SLR(1) parsing tables using FOLLOW sets"""
        print("\n=== BUILDING SLR(1) PARSING TABLES ===\n")
        
        # Compute FOLLOW sets
        follow = self.compute_follow()
        
        slr_action_table = defaultdict(dict)
        slr_goto_table = defaultdict(dict)
        slr_conflicts = []
        
        for state in self.states:
            state_actions = defaultdict(set)
            
            for item in state.items:
                if item.is_complete():
                    if item.lhs == self.augmented_start:
                        state_actions['$'].add('accept')
                    else:
                        prod_num = self.get_production_number(item.lhs, item.rhs)
                        reduce_action = f"r{prod_num}"
                        
                        # SLR(1): Only add reduce for terminals in FOLLOW(A)
                        for terminal in follow.get(item.lhs, set()):
                            state_actions[terminal].add(reduce_action)
                else:
                    next_sym = item.next_symbol()
                    if next_sym in self.terminals:
                        if (state.id, next_sym) in self.transitions:
                            next_state = self.transitions[(state.id, next_sym)]
                            shift_action = f"s{next_state}"
                            state_actions[next_sym].add(shift_action)
            
            for terminal, actions in state_actions.items():
                actions_list = sorted(list(actions))
                
                if len(actions_list) == 1:
                    slr_action_table[state.id][terminal] = actions_list[0]
                else:
                    conflict_type = 'reduce-reduce' if all(a.startswith('r') or a == 'accept' for a in actions_list) else 'shift-reduce'
                    
                    slr_conflicts.append({
                        'state': state.id,
                        'symbol': terminal,
                        'type': conflict_type,
                        'action1': actions_list[0],
                        'action2': actions_list[1] if len(actions_list) > 1 else ''
                    })
                    
                    slr_action_table[state.id][terminal] = ' / '.join(actions_list)
            
            for non_terminal in self.non_terminals:
                if (state.id, non_terminal) in self.transitions:
                    next_state = self.transitions[(state.id, non_terminal)]
                    slr_goto_table[state.id][non_terminal] = next_state
        
        print(f"SLR(1) ACTION table: {len(slr_action_table)} states")
        print(f"SLR(1) GOTO table: {len(slr_goto_table)} states")
        
        if slr_conflicts:
            print(f"\n⚠️  Found {len(slr_conflicts)} SLR(1) conflict(s)!")
        else:
            print("\n✓ No SLR(1) conflicts found!")
        
        return slr_action_table, slr_goto_table, slr_conflicts, follow
    
    def is_slr1(self, slr_conflicts):
        """Check if grammar is SLR(1)"""
        return len(slr_conflicts) == 0
    
    def generate_excel(self, filename='LR0_minimal/lr0_analysis.xlsx', slr_data=None):
        """Generate Excel file with complete analysis"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        self.create_result_sheet(wb, slr_data)
        self.create_grammar_sheet(wb)
        self.create_states_sheet(wb)
        self.create_action_table_sheet(wb)
        self.create_goto_table_sheet(wb)
        
        if slr_data:
            slr_action, slr_goto, slr_conflicts, follow = slr_data
            self.create_slr_action_table_sheet(wb, slr_action, slr_conflicts)
            self.create_slr_goto_table_sheet(wb, slr_goto)
            self.create_follow_sheet(wb, follow)
            self.create_comparison_sheet(wb, slr_conflicts)
        
        self.create_transitions_sheet(wb)
        self.create_hand_written_guide_sheet(wb)
        
        wb.save(filename)
        print(f"\n✅ Excel file created: {filename}")
    
    def create_result_sheet(self, wb, slr_data=None):
        """Create result sheet"""
        ws = wb.create_sheet("Result", 0)
        
        ws['A1'] = "LR(0) and SLR(1) Parser Analysis Result"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:E1')
        
        # LR(0) Results
        ws['A3'] = "LR(0) Analysis"
        ws['A3'].font = Font(bold=True, size=14, color="4472C4")
        ws.merge_cells('A3:B3')
        
        ws['A4'] = "Is LR(0)?"
        ws['A4'].font = Font(bold=True)
        
        is_lr0 = self.is_lr0()
        ws['B4'] = "YES ✓" if is_lr0 else "NO ✗"
        ws['B4'].font = Font(bold=True, size=12, color="008000" if is_lr0 else "FF0000")
        
        ws['A5'] = "Number of States:"
        ws['A5'].font = Font(bold=True)
        ws['B5'] = len(self.states)
        
        ws['A6'] = "LR(0) Conflicts:"
        ws['A6'].font = Font(bold=True)
        ws['B6'] = len(self.conflicts)
        
        # SLR(1) Results
        if slr_data:
            slr_action, slr_goto, slr_conflicts, follow = slr_data
            
            ws['A8'] = "SLR(1) Analysis"
            ws['A8'].font = Font(bold=True, size=14, color="70AD47")
            ws.merge_cells('A8:B8')
            
            ws['A9'] = "Is SLR(1)?"
            ws['A9'].font = Font(bold=True)
            
            is_slr1 = self.is_slr1(slr_conflicts)
            ws['B9'] = "YES ✓" if is_slr1 else "NO ✗"
            ws['B9'].font = Font(bold=True, size=12, color="008000" if is_slr1 else "FF0000")
            
            ws['A10'] = "SLR(1) Conflicts:"
            ws['A10'].font = Font(bold=True)
            ws['B10'] = len(slr_conflicts)
            
            ws['A11'] = "Conflicts Resolved:"
            ws['A11'].font = Font(bold=True)
            ws['B11'] = len(self.conflicts) - len(slr_conflicts)
            ws['B11'].font = Font(color="008000" if len(slr_conflicts) < len(self.conflicts) else "000000")
        
        if self.conflicts:
            ws['A8'] = "Conflicts Details:"
            ws['A8'].font = Font(bold=True, size=12, color="FF0000")
            ws.merge_cells('A8:E8')
            
            ws['A10'] = "State"
            ws['B10'] = "Symbol"
            ws['C10'] = "Conflict Type"
            ws['D10'] = "Action 1"
            ws['E10'] = "Action 2"
            
            for cell in ['A10', 'B10', 'C10', 'D10', 'E10']:
                ws[cell].font = Font(bold=True, size=11)
                ws[cell].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                ws[cell].alignment = Alignment(horizontal='center', vertical='center')
            
            row = 11
            for conflict in self.conflicts:
                ws[f'A{row}'] = conflict['state']
                ws[f'B{row}'] = conflict['symbol']
                ws[f'C{row}'] = conflict['type']
                ws[f'D{row}'] = conflict['action1']
                ws[f'E{row}'] = conflict['action2']
                
                for col in ['A', 'B', 'C', 'D', 'E']:
                    ws[f'{col}{row}'].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                
                row += 1
        else:
            ws['A8'] = "✓ No conflicts found - Grammar is LR(0)!"
            ws['A8'].font = Font(bold=True, size=12, color="008000")
            ws.merge_cells('A8:D8')
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
    
    def create_grammar_sheet(self, wb):
        """Create grammar sheet with correct production numbering"""
        ws = wb.create_sheet("Grammar")
        
        ws['A1'] = "Production #"
        ws['B1'] = "Non-Terminal"
        ws['C1'] = "Production"
        
        for cell in ['A1', 'B1', 'C1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
        
        row = 2
        prod_num = 0
        
        # First, add augmented production (production 0)
        for prod in self.grammar[self.augmented_start]:
            ws[f'A{row}'] = prod_num
            ws[f'B{row}'] = self.augmented_start
            ws[f'C{row}'] = f"{self.augmented_start} → {' '.join(prod)}"
            ws[f'A{row}'].fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            ws[f'B{row}'].fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            ws[f'C{row}'].fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            row += 1
            prod_num += 1
        
        # Then add the rest in sorted order
        for nt in sorted(self.grammar.keys()):
            if nt == self.augmented_start:
                continue
            for prod in self.grammar[nt]:
                ws[f'A{row}'] = prod_num
                ws[f'B{row}'] = nt
                ws[f'C{row}'] = f"{nt} → {' '.join(prod)}"
                row += 1
                prod_num += 1
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
    
    def create_states_sheet(self, wb):
        """Create states sheet"""
        ws = wb.create_sheet("States")
        
        ws['A1'] = "State"
        ws['B1'] = "Items"
        
        for cell in ['A1', 'B1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
        
        row = 2
        for state in self.states:
            ws[f'A{row}'] = f"State {state.id}"
            ws[f'A{row}'].font = Font(bold=True)
            
            items_text = "\n".join(str(item) for item in sorted(state.items, key=str))
            ws[f'B{row}'] = items_text
            ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            
            ws.row_dimensions[row].height = 15 * len(state.items)
            
            row += 1
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 50
    
    def create_action_table_sheet(self, wb):
        """Create ACTION table sheet"""
        ws = wb.create_sheet("ACTION Table")
        
        terminals_sorted = sorted(self.terminals)
        
        ws['A1'] = "State"
        for col_idx, terminal in enumerate(terminals_sorted, start=2):
            ws.cell(1, col_idx, terminal)
        
        for col_idx in range(1, len(terminals_sorted) + 2):
            cell = ws.cell(1, col_idx)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for state in self.states:
            row = state.id + 2
            ws.cell(row, 1, state.id)
            ws.cell(row, 1).font = Font(bold=True)
            
            for col_idx, terminal in enumerate(terminals_sorted, start=2):
                if terminal in self.action_table[state.id]:
                    action = self.action_table[state.id][terminal]
                    ws.cell(row, col_idx, action)
                    
                    if '/' in action:
                        ws.cell(row, col_idx).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        ws.cell(row, col_idx).font = Font(color="FFFFFF", bold=True)
                else:
                    ws.cell(row, col_idx, "-")
                    ws.cell(row, col_idx).alignment = Alignment(horizontal='center')
        
        ws.column_dimensions['A'].width = 10
        for col_idx in range(2, len(terminals_sorted) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12
    
    def create_goto_table_sheet(self, wb):
        """Create GOTO table sheet"""
        ws = wb.create_sheet("GOTO Table")
        
        non_terminals_sorted = sorted(self.non_terminals)
        
        ws['A1'] = "State"
        for col_idx, non_terminal in enumerate(non_terminals_sorted, start=2):
            ws.cell(1, col_idx, non_terminal)
        
        for col_idx in range(1, len(non_terminals_sorted) + 2):
            cell = ws.cell(1, col_idx)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for state in self.states:
            row = state.id + 2
            ws.cell(row, 1, state.id)
            ws.cell(row, 1).font = Font(bold=True)
            
            for col_idx, non_terminal in enumerate(non_terminals_sorted, start=2):
                if non_terminal in self.goto_table[state.id]:
                    goto_state = self.goto_table[state.id][non_terminal]
                    ws.cell(row, col_idx, goto_state)
                else:
                    ws.cell(row, col_idx, "-")
                    ws.cell(row, col_idx).alignment = Alignment(horizontal='center')
        
        ws.column_dimensions['A'].width = 10
        for col_idx in range(2, len(non_terminals_sorted) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12
    
    def create_transitions_sheet(self, wb):
        """Create transitions sheet"""
        ws = wb.create_sheet("Transitions")
        
        ws['A1'] = "From State"
        ws['B1'] = "Symbol"
        ws['C1'] = "To State"
        
        for cell in ['A1', 'B1', 'C1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
        
        row = 2
        for (from_state, symbol), to_state in sorted(self.transitions.items()):
            ws[f'A{row}'] = from_state
            ws[f'B{row}'] = symbol
            ws[f'C{row}'] = to_state
            row += 1
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
    
    def create_slr_action_table_sheet(self, wb, slr_action, slr_conflicts):
        """Create SLR(1) ACTION table sheet"""
        ws = wb.create_sheet("SLR(1) ACTION Table")
        
        terminals_sorted = sorted(self.terminals)
        
        ws['A1'] = "State"
        for col_idx, terminal in enumerate(terminals_sorted, start=2):
            ws.cell(1, col_idx, terminal)
        
        for col_idx in range(1, len(terminals_sorted) + 2):
            cell = ws.cell(1, col_idx)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for state in self.states:
            row = state.id + 2
            ws.cell(row, 1, state.id)
            ws.cell(row, 1).font = Font(bold=True)
            
            for col_idx, terminal in enumerate(terminals_sorted, start=2):
                if terminal in slr_action[state.id]:
                    action = slr_action[state.id][terminal]
                    ws.cell(row, col_idx, action)
                    
                    if '/' in action:
                        ws.cell(row, col_idx).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        ws.cell(row, col_idx).font = Font(color="FFFFFF", bold=True)
                else:
                    ws.cell(row, col_idx, "-")
                    ws.cell(row, col_idx).alignment = Alignment(horizontal='center')
        
        ws.column_dimensions['A'].width = 10
        for col_idx in range(2, len(terminals_sorted) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12
    
    def create_slr_goto_table_sheet(self, wb, slr_goto):
        """Create SLR(1) GOTO table sheet"""
        ws = wb.create_sheet("SLR(1) GOTO Table")
        
        non_terminals_sorted = sorted(self.non_terminals)
        
        ws['A1'] = "State"
        for col_idx, non_terminal in enumerate(non_terminals_sorted, start=2):
            ws.cell(1, col_idx, non_terminal)
        
        for col_idx in range(1, len(non_terminals_sorted) + 2):
            cell = ws.cell(1, col_idx)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for state in self.states:
            row = state.id + 2
            ws.cell(row, 1, state.id)
            ws.cell(row, 1).font = Font(bold=True)
            
            for col_idx, non_terminal in enumerate(non_terminals_sorted, start=2):
                if non_terminal in slr_goto[state.id]:
                    goto_state = slr_goto[state.id][non_terminal]
                    ws.cell(row, col_idx, goto_state)
                else:
                    ws.cell(row, col_idx, "-")
                    ws.cell(row, col_idx).alignment = Alignment(horizontal='center')
        
        ws.column_dimensions['A'].width = 10
        for col_idx in range(2, len(non_terminals_sorted) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12
    
    def create_follow_sheet(self, wb, follow):
        """Create FOLLOW sets sheet"""
        ws = wb.create_sheet("FOLLOW Sets")
        
        ws['A1'] = "Non-Terminal"
        ws['B1'] = "FOLLOW Set"
        
        for cell in ['A1', 'B1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
        
        row = 2
        for A in sorted(self.non_terminals):
            ws[f'A{row}'] = A
            ws[f'B{row}'] = ', '.join(sorted(follow.get(A, set())))
            row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
    
    def create_comparison_sheet(self, wb, slr_conflicts):
        """Create comparison sheet between LR(0) and SLR(1)"""
        ws = wb.create_sheet("LR(0) vs SLR(1)")
        
        ws['A1'] = "Comparison: LR(0) vs SLR(1)"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        ws['A3'] = "Metric"
        ws['B3'] = "LR(0)"
        ws['C3'] = "SLR(1)"
        ws['D3'] = "Improvement"
        
        for cell in ['A3', 'B3', 'C3', 'D3']:
            ws[cell].font = Font(bold=True, size=11)
            ws[cell].fill = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
            ws[cell].font = Font(bold=True, size=11, color="FFFFFF")
        
        row = 4
        
        ws[f'A{row}'] = "Conflicts"
        ws[f'B{row}'] = len(self.conflicts)
        ws[f'C{row}'] = len(slr_conflicts)
        ws[f'D{row}'] = len(self.conflicts) - len(slr_conflicts)
        ws[f'D{row}'].font = Font(color="008000" if len(slr_conflicts) < len(self.conflicts) else "000000")
        row += 1
        
        ws[f'A{row}'] = "Is Accepted?"
        ws[f'B{row}'] = "NO" if len(self.conflicts) > 0 else "YES"
        ws[f'C{row}'] = "NO" if len(slr_conflicts) > 0 else "YES"
        ws[f'D{row}'] = "✓ Resolved!" if len(slr_conflicts) == 0 and len(self.conflicts) > 0 else "-"
        ws[f'D{row}'].font = Font(color="008000" if len(slr_conflicts) == 0 and len(self.conflicts) > 0 else "000000")
        row += 1
        
        ws[f'A{row}'] = "Reduce Strategy"
        ws[f'B{row}'] = "All terminals"
        ws[f'C{row}'] = "FOLLOW set only"
        ws[f'D{row}'] = "More precise"
        row += 2
        
        ws[f'A{row}'] = "Conflicts Resolved:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        resolved = []
        for lr0_conf in self.conflicts:
            found = False
            for slr_conf in slr_conflicts:
                if (lr0_conf['state'] == slr_conf['state'] and 
                    lr0_conf['symbol'] == slr_conf['symbol']):
                    found = True
                    break
            if not found:
                resolved.append(lr0_conf)
        
        if resolved:
            for conf in resolved:
                ws[f'A{row}'] = f"State {conf['state']}, Symbol '{conf['symbol']}'"
                ws[f'A{row}'].font = Font(color="008000")
                row += 1
        else:
            ws[f'A{row}'] = "None" if len(self.conflicts) == 0 else "No conflicts resolved"
            row += 1
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
    
    def create_hand_written_guide_sheet(self, wb):
        """Create guide for hand-written parsing tables"""
        ws = wb.create_sheet("Hand-Written Guide")
        
        ws['A1'] = "Step-by-Step Guide for Hand-Written LR(0) and SLR(1) Tables"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        row = 3
        ws[f'A{row}'] = "LR(0) Steps:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        steps = [
            ("Step 1: Augment Grammar", "Add S' → S production"),
            ("Step 2: Build LR(0) Items", "Create items with dot positions"),
            ("Step 3: Compute Closures", "Add items for non-terminals after dot"),
            ("Step 4: Build States", "Use GOTO to create new states"),
            ("Step 5: Build ACTION Table", "Add shift and reduce actions (ALL terminals)"),
            ("Step 6: Build GOTO Table", "Add state transitions for non-terminals"),
            ("Step 7: Check Conflicts", "Look for cells with multiple actions"),
        ]
        
        for step, description in steps:
            ws[f'A{row}'] = step
            ws[f'A{row}'].font = Font(bold=True, size=10)
            ws[f'B{row}'] = description
            row += 1
        
        row += 2
        ws[f'A{row}'] = "SLR(1) Steps:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        slr_steps = [
            ("Step 1-4: Same as LR(0)", "Build same automaton"),
            ("Step 5: Compute FOLLOW", "Calculate FOLLOW sets for all non-terminals"),
            ("Step 6: Build ACTION Table", "Reduce only for terminals in FOLLOW(A)"),
            ("Step 7: Build GOTO Table", "Same as LR(0)"),
            ("Step 8: Check Conflicts", "Should have fewer conflicts than LR(0)"),
        ]
        
        for step, description in slr_steps:
            ws[f'A{row}'] = step
            ws[f'A{row}'].font = Font(bold=True, size=10)
            ws[f'B{row}'] = description
            row += 1
        
        row += 2
        ws[f'A{row}'] = "Table Dimensions:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = f"States: {len(self.states)}"
        row += 1
        ws[f'A{row}'] = f"ACTION Table: {len(self.states)} × {len(self.terminals)}"
        row += 1
        ws[f'A{row}'] = f"GOTO Table: {len(self.states)} × {len(self.non_terminals)}"
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50


def main():
    from grammar import grammar
    
    print("="*70)
    print("LR(0) AND SLR(1) PARSER ANALYZER - MINIMAL VERSION")
    print("="*70)
    
    analyzer = LR0Analyzer(grammar)
    
    print(f"\nGrammar has {len(analyzer.non_terminals)} non-terminals and {len(analyzer.terminals)} terminals")
    print(f"Start symbol: {analyzer.start_symbol}")
    print(f"Augmented start: {analyzer.augmented_start}")
    
    # Build LR(0) automaton and tables
    analyzer.build_automaton()
    analyzer.build_parsing_tables()
    
    # Build SLR(1) tables
    slr_action, slr_goto, slr_conflicts, follow = analyzer.build_slr_parsing_tables()
    
    # Generate Excel with both analyses
    analyzer.generate_excel(slr_data=(slr_action, slr_goto, slr_conflicts, follow))
    
    print("\n" + "="*70)
    print("LR(0) RESULT")
    print("="*70)
    
    if analyzer.is_lr0():
        print("✅ The grammar IS LR(0)")
    else:
        print("❌ The grammar IS NOT LR(0)")
        print(f"\nFound {len(analyzer.conflicts)} LR(0) conflict(s):")
        for i, conflict in enumerate(analyzer.conflicts, 1):
            print(f"\n  Conflict {i}:")
            print(f"    State: {conflict['state']}")
            print(f"    Symbol: {conflict['symbol']}")
            print(f"    Type: {conflict['type']}")
            print(f"    Action 1: {conflict['action1']}")
            print(f"    Action 2: {conflict['action2']}")
    
    print("\n" + "="*70)
    print("SLR(1) RESULT")
    print("="*70)
    
    if analyzer.is_slr1(slr_conflicts):
        print("✅ The grammar IS SLR(1)")
        print(f"\n🎉 SLR(1) resolved all {len(analyzer.conflicts)} LR(0) conflicts!")
    else:
        print("❌ The grammar IS NOT SLR(1)")
        print(f"\nFound {len(slr_conflicts)} SLR(1) conflict(s):")
        for i, conflict in enumerate(slr_conflicts, 1):
            print(f"\n  Conflict {i}:")
            print(f"    State: {conflict['state']}")
            print(f"    Symbol: {conflict['symbol']}")
            print(f"    Type: {conflict['type']}")
            print(f"    Action 1: {conflict['action1']}")
            print(f"    Action 2: {conflict['action2']}")
        
        resolved = len(analyzer.conflicts) - len(slr_conflicts)
        if resolved > 0:
            print(f"\n✓ SLR(1) resolved {resolved} out of {len(analyzer.conflicts)} conflicts")
    
    print("\n" + "="*70)


if __name__ == "__main__":
    main()

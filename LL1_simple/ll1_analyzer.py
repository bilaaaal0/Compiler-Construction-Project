"""
LL(1) Grammar Analyzer
Performs left factoring, left recursion elimination, and computes FIRST, FOLLOW, NULLABLE sets
Creates predictive parsing table and determines if grammar is LL(1)
"""

from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class LL1Analyzer:
    def __init__(self, grammar_text):
        self.original_grammar = {}
        self.grammar = {}
        self.terminals = set()
        self.non_terminals = set()
        self.nullable = set()
        self.first = defaultdict(set)
        self.follow = defaultdict(set)
        self.parsing_table = defaultdict(dict)
        self.conflicts = []
        self.parse_grammar(grammar_text)
        
    def parse_grammar(self, grammar_text):
        """Parse the grammar from text format"""
        lines = grammar_text.strip().split('\n')
        current_lhs = None
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
                
            if '→' in line:
                parts = line.split('→')
                lhs = parts[0].strip()
                rhs = parts[1].strip()
                current_lhs = lhs
                self.non_terminals.add(lhs)
                
                if lhs not in self.original_grammar:
                    self.original_grammar[lhs] = []
                self.original_grammar[lhs].append(self.parse_production(rhs))
                
            elif '|' in line:
                rhs = line.split('|')[1].strip()
                if current_lhs:
                    self.original_grammar[current_lhs].append(self.parse_production(rhs))
        
        # Extract terminals
        for lhs, productions in self.original_grammar.items():
            for prod in productions:
                for symbol in prod:
                    if symbol not in self.non_terminals and symbol != 'ε':
                        self.terminals.add(symbol)
        
        # Add end marker
        self.terminals.add('$')
        
        # Copy to working grammar
        self.grammar = {k: [p[:] for p in v] for k, v in self.original_grammar.items()}
    
    def parse_production(self, rhs):
        """Parse a production right-hand side into symbols"""
        if rhs == 'ε':
            return ['ε']
        
        symbols = []
        i = 0
        while i < len(rhs):
            if rhs[i] == "'":
                # Terminal in quotes
                j = i + 1
                while j < len(rhs) and rhs[j] != "'":
                    j += 1
                symbols.append(rhs[i:j+1])
                i = j + 1
            elif rhs[i].isspace():
                i += 1
            else:
                # Non-terminal or identifier
                j = i
                while j < len(rhs) and not rhs[j].isspace() and rhs[j] != "'":
                    j += 1
                symbol = rhs[i:j]
                if symbol:
                    symbols.append(symbol)
                i = j
        
        return symbols
    
    def eliminate_left_recursion(self):
        """Eliminate left recursion from grammar"""
        print("\n=== ELIMINATING LEFT RECURSION ===\n")
        
        non_terminals_list = list(self.grammar.keys())
        new_grammar = {}
        changes_made = False
        
        for i, A in enumerate(non_terminals_list):
            # Eliminate indirect left recursion
            for j in range(i):
                B = non_terminals_list[j]
                new_productions = []
                
                for production in self.grammar[A]:
                    if production[0] == B:
                        # Replace B with its productions
                        for b_prod in self.grammar[B]:
                            new_prod = b_prod + production[1:]
                            new_productions.append(new_prod)
                        changes_made = True
                    else:
                        new_productions.append(production)
                
                self.grammar[A] = new_productions
            
            # Eliminate direct left recursion
            alpha_prods = []  # Productions with left recursion
            beta_prods = []   # Productions without left recursion
            
            for production in self.grammar[A]:
                if production[0] == A:
                    alpha_prods.append(production[1:])
                else:
                    beta_prods.append(production)
            
            if alpha_prods:
                # Create new non-terminal A'
                A_prime = A + "'"
                self.non_terminals.add(A_prime)
                
                # A → β A'
                new_grammar[A] = []
                for beta in beta_prods:
                    new_grammar[A].append(beta + [A_prime])
                
                # A' → α A' | ε
                new_grammar[A_prime] = []
                for alpha in alpha_prods:
                    new_grammar[A_prime].append(alpha + [A_prime])
                new_grammar[A_prime].append(['ε'])
                
                print(f"Eliminated left recursion in {A}:")
                print(f"  {A} → {' | '.join([' '.join(p) for p in new_grammar[A]])}")
                print(f"  {A_prime} → {' | '.join([' '.join(p) for p in new_grammar[A_prime]])}")
                changes_made = True
            else:
                new_grammar[A] = self.grammar[A]
        
        self.grammar = new_grammar
        
        if not changes_made:
            print("No left recursion found.")
        
        return changes_made
    
    def left_factor(self):
        """Perform left factoring on grammar"""
        print("\n=== PERFORMING LEFT FACTORING ===\n")
        
        changes_made = True
        iteration = 0
        
        while changes_made:
            changes_made = False
            iteration += 1
            new_grammar = {}
            
            for A in list(self.grammar.keys()):
                productions = self.grammar[A]
                
                # Group productions by common prefix
                prefix_groups = defaultdict(list)
                
                for prod in productions:
                    if prod != ['ε']:
                        prefix_groups[prod[0]].append(prod)
                    else:
                        prefix_groups['ε'].append(prod)
                
                # Check for common prefixes
                factored = False
                for prefix, prods in prefix_groups.items():
                    if len(prods) > 1 and prefix != 'ε':
                        # Find longest common prefix
                        common_prefix = [prefix]
                        min_len = min(len(p) for p in prods)
                        
                        for i in range(1, min_len):
                            if all(p[i] == prods[0][i] for p in prods):
                                common_prefix.append(prods[0][i])
                            else:
                                break
                        
                        if len(common_prefix) > 0:
                            # Create new non-terminal
                            A_prime = A + "'"
                            counter = 1
                            while A_prime in self.non_terminals:
                                A_prime = A + "'" * (counter + 1)
                                counter += 1
                            
                            self.non_terminals.add(A_prime)
                            
                            # A → α A'
                            new_grammar[A] = []
                            new_grammar[A].append(common_prefix + [A_prime])
                            
                            # Add other productions that don't share prefix
                            for prod in productions:
                                if prod not in prods:
                                    new_grammar[A].append(prod)
                            
                            # A' → β1 | β2 | ...
                            new_grammar[A_prime] = []
                            for prod in prods:
                                suffix = prod[len(common_prefix):]
                                if not suffix:
                                    suffix = ['ε']
                                new_grammar[A_prime].append(suffix)
                            
                            print(f"Left factored {A} (iteration {iteration}):")
                            print(f"  Common prefix: {' '.join(common_prefix)}")
                            print(f"  {A} → {' | '.join([' '.join(p) for p in new_grammar[A]])}")
                            print(f"  {A_prime} → {' | '.join([' '.join(p) for p in new_grammar[A_prime]])}")
                            
                            factored = True
                            changes_made = True
                            break
                
                if not factored:
                    new_grammar[A] = productions
            
            self.grammar = new_grammar
        
        if iteration == 1:
            print("No left factoring needed.")
    
    def compute_nullable(self):
        """Compute NULLABLE set"""
        print("\n=== COMPUTING NULLABLE ===\n")
        
        changed = True
        while changed:
            changed = False
            for A in self.non_terminals:
                if A in self.nullable:
                    continue
                
                for production in self.grammar.get(A, []):
                    if production == ['ε']:
                        self.nullable.add(A)
                        changed = True
                        break
                    
                    if all(symbol in self.nullable for symbol in production if symbol in self.non_terminals):
                        if all(symbol in self.non_terminals for symbol in production):
                            self.nullable.add(A)
                            changed = True
                            break
        
        print(f"NULLABLE = {{{', '.join(sorted(self.nullable))}}}")
    
    def compute_first(self):
        """Compute FIRST sets"""
        print("\n=== COMPUTING FIRST SETS ===\n")
        
        # Initialize FIRST for terminals
        for t in self.terminals:
            self.first[t].add(t)
        
        # Compute FIRST for non-terminals
        changed = True
        while changed:
            changed = False
            for A in self.non_terminals:
                for production in self.grammar.get(A, []):
                    if production == ['ε']:
                        if 'ε' not in self.first[A]:
                            self.first[A].add('ε')
                            changed = True
                    else:
                        for symbol in production:
                            # Add FIRST(symbol) - {ε} to FIRST(A)
                            before = len(self.first[A])
                            self.first[A].update(self.first[symbol] - {'ε'})
                            if len(self.first[A]) > before:
                                changed = True
                            
                            # If symbol is not nullable, stop
                            if symbol not in self.nullable:
                                break
                        else:
                            # All symbols are nullable
                            if 'ε' not in self.first[A]:
                                self.first[A].add('ε')
                                changed = True
        
        for A in sorted(self.non_terminals):
            print(f"FIRST({A}) = {{{', '.join(sorted(self.first[A]))}}}")
    
    def compute_follow(self):
        """Compute FOLLOW sets"""
        print("\n=== COMPUTING FOLLOW SETS ===\n")
        
        # Start symbol gets $
        start_symbol = list(self.grammar.keys())[0]
        self.follow[start_symbol].add('$')
        
        changed = True
        while changed:
            changed = False
            for A in self.non_terminals:
                for production in self.grammar.get(A, []):
                    for i, symbol in enumerate(production):
                        if symbol in self.non_terminals:
                            # Add FIRST(β) - {ε} to FOLLOW(symbol)
                            beta = production[i+1:]
                            
                            if beta:
                                for b_symbol in beta:
                                    before = len(self.follow[symbol])
                                    self.follow[symbol].update(self.first[b_symbol] - {'ε'})
                                    if len(self.follow[symbol]) > before:
                                        changed = True
                                    
                                    if b_symbol not in self.nullable:
                                        break
                                else:
                                    # All of β is nullable
                                    before = len(self.follow[symbol])
                                    self.follow[symbol].update(self.follow[A])
                                    if len(self.follow[symbol]) > before:
                                        changed = True
                            else:
                                # symbol is last
                                before = len(self.follow[symbol])
                                self.follow[symbol].update(self.follow[A])
                                if len(self.follow[symbol]) > before:
                                    changed = True
        
        for A in sorted(self.non_terminals):
            print(f"FOLLOW({A}) = {{{', '.join(sorted(self.follow[A]))}}}")
    
    def compute_first_of_string(self, symbols):
        """Compute FIRST of a string of symbols"""
        result = set()
        
        for symbol in symbols:
            result.update(self.first[symbol] - {'ε'})
            if symbol not in self.nullable:
                break
        else:
            result.add('ε')
        
        return result
    
    def build_parsing_table(self):
        """Build LL(1) predictive parsing table"""
        print("\n=== BUILDING PARSING TABLE ===\n")
        
        for A in self.non_terminals:
            for production in self.grammar.get(A, []):
                # Compute FIRST(α)
                first_alpha = self.compute_first_of_string(production)
                
                # For each terminal in FIRST(α), add A → α to table
                for terminal in first_alpha - {'ε'}:
                    if terminal in self.parsing_table[A]:
                        # Conflict!
                        self.conflicts.append({
                            'non_terminal': A,
                            'terminal': terminal,
                            'production1': self.parsing_table[A][terminal],
                            'production2': production
                        })
                        # Use format: S→A / S→B
                        existing = ' '.join(self.parsing_table[A][terminal])
                        new = ' '.join(production)
                        self.parsing_table[A][terminal] = [f"{A}→{existing} / {A}→{new}"]
                    else:
                        self.parsing_table[A][terminal] = production
                
                # If ε in FIRST(α), add A → α for each terminal in FOLLOW(A)
                if 'ε' in first_alpha:
                    for terminal in self.follow[A]:
                        if terminal in self.parsing_table[A]:
                            # Conflict!
                            self.conflicts.append({
                                'non_terminal': A,
                                'terminal': terminal,
                                'production1': self.parsing_table[A][terminal],
                                'production2': production
                            })
                            existing = ' '.join(self.parsing_table[A][terminal])
                            new = ' '.join(production)
                            self.parsing_table[A][terminal] = [f"{A}→{existing} / {A}→{new}"]
                        else:
                            self.parsing_table[A][terminal] = production
        
        print("Parsing table built successfully.")
        if self.conflicts:
            print(f"\n⚠️  Found {len(self.conflicts)} conflict(s)!")
    
    def is_ll1(self):
        """Check if grammar is LL(1)"""
        return len(self.conflicts) == 0
    
    def generate_excel(self, filename='LL1_simple/grammar_analysis.xlsx'):
        """Generate Excel file with complete analysis"""
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        self.create_original_grammar_sheet(wb)
        self.create_transformed_grammar_sheet(wb)
        self.create_nullable_sheet(wb)
        self.create_first_sheet(wb)
        self.create_follow_sheet(wb)
        self.create_parsing_table_sheet(wb)
        self.create_result_sheet(wb)
        
        wb.save(filename)
        print(f"\n✅ Excel file created: {filename}")
    
    def create_original_grammar_sheet(self, wb):
        """Create sheet with original grammar"""
        ws = wb.create_sheet("Original Grammar")
        
        # Headers
        ws['A1'] = "Non-Terminal"
        ws['B1'] = "Productions"
        
        # Style headers
        for cell in ['A1', 'B1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
        
        row = 2
        for A in self.original_grammar.keys():
            ws[f'A{row}'] = A
            productions = ' | '.join([' '.join(p) for p in self.original_grammar[A]])
            ws[f'B{row}'] = productions
            row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 80
    
    def create_transformed_grammar_sheet(self, wb):
        """Create sheet with transformed grammar"""
        ws = wb.create_sheet("Transformed Grammar")
        
        # Headers
        ws['A1'] = "Non-Terminal"
        ws['B1'] = "Productions"
        
        # Style headers
        for cell in ['A1', 'B1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
        
        row = 2
        for A in self.grammar.keys():
            ws[f'A{row}'] = A
            productions = ' | '.join([' '.join(p) for p in self.grammar[A]])
            ws[f'B{row}'] = productions
            row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 80
    
    def create_nullable_sheet(self, wb):
        """Create sheet with NULLABLE set"""
        ws = wb.create_sheet("NULLABLE")
        
        ws['A1'] = "NULLABLE Non-Terminals"
        ws['A1'].font = Font(bold=True, size=12)
        ws['A1'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        row = 2
        for A in sorted(self.nullable):
            ws[f'A{row}'] = A
            row += 1
        
        ws.column_dimensions['A'].width = 25
    
    def create_first_sheet(self, wb):
        """Create sheet with FIRST sets"""
        ws = wb.create_sheet("FIRST Sets")
        
        ws['A1'] = "Non-Terminal"
        ws['B1'] = "FIRST Set"
        
        for cell in ['A1', 'B1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
        
        row = 2
        for A in sorted(self.non_terminals):
            ws[f'A{row}'] = A
            ws[f'B{row}'] = ', '.join(sorted(self.first[A]))
            row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 60
    
    def create_follow_sheet(self, wb):
        """Create sheet with FOLLOW sets"""
        ws = wb.create_sheet("FOLLOW Sets")
        
        ws['A1'] = "Non-Terminal"
        ws['B1'] = "FOLLOW Set"
        
        for cell in ['A1', 'B1']:
            ws[cell].font = Font(bold=True, size=12)
            ws[cell].fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
            ws[cell].font = Font(bold=True, size=12, color="FFFFFF")
        
        row = 2
        for A in sorted(self.non_terminals):
            ws[f'A{row}'] = A
            ws[f'B{row}'] = ', '.join(sorted(self.follow[A]))
            row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 60
    
    def create_parsing_table_sheet(self, wb):
        """Create sheet with parsing table"""
        ws = wb.create_sheet("Parsing Table")
        
        # Get sorted terminals and non-terminals
        terminals_sorted = sorted(self.terminals)
        non_terminals_sorted = sorted(self.non_terminals)
        
        # Headers
        ws['A1'] = "Non-Terminal"
        for col_idx, terminal in enumerate(terminals_sorted, start=2):
            ws.cell(1, col_idx, terminal)
        
        # Style headers
        for col_idx in range(1, len(terminals_sorted) + 2):
            cell = ws.cell(1, col_idx)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Fill table
        for row_idx, non_terminal in enumerate(non_terminals_sorted, start=2):
            ws.cell(row_idx, 1, non_terminal)
            ws.cell(row_idx, 1).font = Font(bold=True)
            
            for col_idx, terminal in enumerate(terminals_sorted, start=2):
                if terminal in self.parsing_table[non_terminal]:
                    production = self.parsing_table[non_terminal][terminal]
                    prod_str = ' '.join(production)
                    
                    # Check if it's a conflict (contains '/')
                    if '/' in prod_str:
                        ws.cell(row_idx, col_idx, prod_str)
                        ws.cell(row_idx, col_idx).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        ws.cell(row_idx, col_idx).font = Font(color="FFFFFF", bold=True)
                    else:
                        ws.cell(row_idx, col_idx, f"{non_terminal} → {prod_str}")
                else:
                    # Empty cell - add dash
                    ws.cell(row_idx, col_idx, "-")
                    ws.cell(row_idx, col_idx).alignment = Alignment(horizontal='center', vertical='center')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        for col_idx in range(2, len(terminals_sorted) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 25
    
    def create_result_sheet(self, wb):
        """Create sheet with final result"""
        ws = wb.create_sheet("Result", 0)  # Make it first sheet
        
        ws['A1'] = "LL(1) Grammar Analysis Result"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        ws['A3'] = "Is LL(1)?"
        ws['A3'].font = Font(bold=True, size=14)
        
        is_ll1 = self.is_ll1()
        ws['B3'] = "YES ✓" if is_ll1 else "NO ✗"
        ws['B3'].font = Font(bold=True, size=14, color="008000" if is_ll1 else "FF0000")
        
        ws['A5'] = "Number of Conflicts:"
        ws['A5'].font = Font(bold=True)
        ws['B5'] = len(self.conflicts)
        
        if self.conflicts:
            ws['A7'] = "Conflicts Details:"
            ws['A7'].font = Font(bold=True, size=12, color="FF0000")
            ws.merge_cells('A7:D7')
            
            # Headers for conflict table
            ws['A9'] = "Non-Terminal"
            ws['B9'] = "Terminal"
            ws['C9'] = "Production 1"
            ws['D9'] = "Production 2"
            
            for cell in ['A9', 'B9', 'C9', 'D9']:
                ws[cell].font = Font(bold=True, size=11)
                ws[cell].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                ws[cell].alignment = Alignment(horizontal='center', vertical='center')
            
            row = 10
            for i, conflict in enumerate(self.conflicts, 1):
                ws[f'A{row}'] = conflict['non_terminal']
                ws[f'B{row}'] = conflict['terminal']
                
                prod1 = ' '.join(conflict['production1'])
                prod2 = ' '.join(conflict['production2'])
                
                ws[f'C{row}'] = f"{conflict['non_terminal']} → {prod1}"
                ws[f'D{row}'] = f"{conflict['non_terminal']} → {prod2}"
                
                # Highlight conflict rows
                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{row}'].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                
                row += 1
        else:
            ws['A7'] = "✓ No conflicts found - Grammar is LL(1)!"
            ws['A7'].font = Font(bold=True, size=12, color="008000")
            ws.merge_cells('A7:D7')
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 40


def main():
    from grammar import grammar
    
    print("="*70)
    print("LL(1) GRAMMAR ANALYZER")
    print("="*70)
    
    analyzer = LL1Analyzer(grammar)
    
    print(f"\nOriginal Grammar has {len(analyzer.non_terminals)} non-terminals and {len(analyzer.terminals)} terminals")
    
    # Perform transformations
    analyzer.eliminate_left_recursion()
    analyzer.left_factor()
    
    # Compute sets
    analyzer.compute_nullable()
    analyzer.compute_first()
    analyzer.compute_follow()
    
    # Build parsing table
    analyzer.build_parsing_table()
    
    # Generate Excel
    analyzer.generate_excel()
    
    # Final result
    print("\n" + "="*70)
    print("FINAL RESULT")
    print("="*70)
    
    if analyzer.is_ll1():
        print("✅ The grammar IS LL(1)")
    else:
        print("❌ The grammar IS NOT LL(1)")
        print(f"\nFound {len(analyzer.conflicts)} conflict(s):")
        for i, conflict in enumerate(analyzer.conflicts, 1):
            print(f"\n  Conflict {i}:")
            print(f"    Non-terminal: {conflict['non_terminal']}")
            print(f"    Terminal: {conflict['terminal']}")
            print(f"    Production 1: {' '.join(conflict['production1'])}")
            print(f"    Production 2: {' '.join(conflict['production2'])}")
    
    print("\n" + "="*70)


if __name__ == "__main__":
    main()
